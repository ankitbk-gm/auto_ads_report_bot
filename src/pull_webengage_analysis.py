"""
pull_webengage_analysis.py
Builds the WebEngage Analysis report with 5 sections.

Google Sheet: all tables
Excel: Report 1,2,5 unchanged | Journey classified only | Push no segment table

Reads from: WebEngage sheet (WEBENGAGE_SHEET_ID)
Writes to:  Report sheet (REPORT_SHEET_ID): Webengage_Analysis tab
            OneDrive Excel: Team Wise Cost.xlsx -> Webengage_Analysis sheet
Period: MTD (Reporting Period Start Date in current month)
"""

import os
import re
from datetime import datetime
from collections import defaultdict
import pytz
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials

load_dotenv()

IST = pytz.timezone("Asia/Kolkata")
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

WEBENGAGE_SHEET_ID = os.getenv("WEBENGAGE_SHEET_ID")
REPORT_SHEET_ID    = os.getenv("REPORT_SHEET_ID")
ONEDRIVE_PATH      = r"C:\Users\anitb\OneDrive - Agrim Wholesale Private Limited\Marketing_Reports\Team Wise Cost.xlsx"
REPORT_TAB         = "Webengage_Analysis"


# ── Auth ───────────────────────────────────────────────────────────────────────

def get_gc():
    creds_path = (
        os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH")
        or os.getenv("GOOGLE_CREDENTIALS_PATH")
        or "service_account.json"
    )
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


def read_sheet(gc, sheet_id, tab_name):
    sh   = gc.open_by_key(sheet_id)
    ws   = sh.worksheet(tab_name)
    rows = ws.get_all_values()
    if not rows or len(rows) < 2:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


# ── MTD filter ─────────────────────────────────────────────────────────────────

def is_current_month(date_str):
    now = datetime.now(IST)
    s   = str(date_str).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            d = datetime.strptime(s[:10], fmt)
            return d.month == now.month and d.year == now.year
        except ValueError:
            continue
    return False


def to_float(val):
    try:
        return float(val or 0)
    except (ValueError, TypeError):
        return 0.0


def safe_pct(num, den):
    if not den:
        return "N/A"
    return f"{round(num / den * 100, 2)}%"


# ── Team mapping ───────────────────────────────────────────────────────────────

def map_team_webengage(campaign_type, campaign_name):
    t = str(campaign_type).strip().lower()
    n = campaign_name.lower().strip()
    if t == "relay":
        return "Marketing Retention"
    if t == "journey":
        return "Marketing Retention" if "add to cart but not purchased" in n else "Marketing Onboarding"
    if t in ("one time", "one-time"):
        return "Marketing Onboarding" if "onboarding" in n else "Marketing Retention"
    return "Marketing Retention"


# ── Push campaign type ─────────────────────────────────────────────────────────

CAMPAIGN_TYPE_MAP = {
    "top_selling_seeds":    "Top Selling Seeds",
    "top_selling_skus":     "Top Selling SKUs",
    "top_branded_products": "Top Branded Products",
    "top_products":         "Top Products",
    "top_selling":          "Top Selling",
    "top_technical":        "TOP TECHNICAL",
    "volume_discount":      "Volume Discount",
    "coupon_discount":      "Coupon Discount",
    "cp_push":              "CP PUSH",
    "under_99":             "Under 99 CP/CN",
    "price_drop":           "Price Drop",
    "back_in_stock":        "Back in Stock",
    "brand":                "Brand",
    "discount":             "Discount",
}


def extract_push_type(campaign_name):
    n = campaign_name.lower().replace(" ", "_").replace("-", "_")
    for key, display in CAMPAIGN_TYPE_MAP.items():
        if key in n:
            return display
    return "Other"


# ── Journey classification ─────────────────────────────────────────────────────

JOURNEY_CLASS_MAP = {
    "first_hpv":  "HPV_Not_View_Content",
    "hpv":        "HPV_Not_View_Content",
    "add_to_cart":"Add_to_Cart",
    "agro_tool":  "Agro_Tools_Cross_Sell",
    "cross_sell": "Agro_Tools_Cross_Sell",
    "first_open": "KYC_Journey",
    "kyc":        "KYC_Journey",
}


def classify_journey(journey_name):
    n = journey_name.lower().replace(" ", "_")
    for key in ["first_hpv", "add_to_cart", "agro_tool", "cross_sell", "first_open", "hpv", "kyc"]:
        if key in n:
            return JOURNEY_CLASS_MAP[key]
    return "Other"


# ══════════════════════════════════════════════════════════════════════════════
# REPORT 1 — Channel Performance
# ══════════════════════════════════════════════════════════════════════════════

def build_channel_performance(gc):
    print("[Report 1] Channel Performance...")
    wa_d = wa_c = wa_v = 0
    for r in read_sheet(gc, WEBENGAGE_SHEET_ID, "Whatsapp_Campaign"):
        if not is_current_month(r.get("Reporting Period Start Date", "")): continue
        wa_d += to_float(r.get("Delivered"))
        wa_c += to_float(r.get("Unique Clicks"))
        wa_v += to_float(r.get("Unique Conversions"))

    pu_i = pu_c = pu_v = 0
    for r in read_sheet(gc, WEBENGAGE_SHEET_ID, "Push_Campaign"):
        if not is_current_month(r.get("Reporting Period Start Date", "")): continue
        pu_i += to_float(r.get("Total Impressions"))
        pu_c += to_float(r.get("Total Clicks"))
        pu_v += to_float(r.get("Unique Conversions"))

    ia_i = ia_c = ia_v = 0
    for r in read_sheet(gc, WEBENGAGE_SHEET_ID, "Inapp_Campaign"):
        if not is_current_month(r.get("Reporting Period Start Date", "")): continue
        ia_i += to_float(r.get("Unique Impressions"))
        ia_c += to_float(r.get("Unique Clicks"))
        ia_v += to_float(r.get("Unique Conversions"))

    headers = ["Channel", "Delivered / Impr", "Clicks", "CTR", "Conversions"]
    rows = [
        ["WhatsApp", int(wa_d), int(wa_c), safe_pct(wa_c, wa_d), int(wa_v)],
        ["Push",     int(pu_i), int(pu_c), safe_pct(pu_c, pu_i), int(pu_v)],
        ["In-app",   int(ia_i), int(ia_c), safe_pct(ia_c, ia_i), int(ia_v)],
    ]
    print(f"  WA={int(wa_d)} Push={int(pu_i)} Inapp={int(ia_i)}")
    return headers, rows


# ══════════════════════════════════════════════════════════════════════════════
# REPORT 2 — Team-wise Engagement
# ══════════════════════════════════════════════════════════════════════════════

def build_team_engagement(gc):
    print("[Report 2] Team Engagement...")
    TEAMS    = ["Marketing Onboarding", "Marketing Retention"]
    CHANNELS = ["WhatsApp", "Push", "In-app"]
    data = defaultdict(lambda: defaultdict(lambda: {"d": 0, "c": 0, "v": 0}))

    for r in read_sheet(gc, WEBENGAGE_SHEET_ID, "Whatsapp_Campaign"):
        if not is_current_month(r.get("Reporting Period Start Date", "")): continue
        team = map_team_webengage(r.get("Type of Campaign", ""), r.get("Campaign Name", ""))
        if team not in TEAMS: continue
        data[team]["WhatsApp"]["d"] += to_float(r.get("Delivered"))
        data[team]["WhatsApp"]["c"] += to_float(r.get("Unique Clicks"))
        data[team]["WhatsApp"]["v"] += to_float(r.get("Unique Conversions"))

    for r in read_sheet(gc, WEBENGAGE_SHEET_ID, "Push_Campaign"):
        if not is_current_month(r.get("Reporting Period Start Date", "")): continue
        team = map_team_webengage(r.get("Type of Campaign", ""), r.get("Campaign Name", ""))
        if team not in TEAMS: continue
        data[team]["Push"]["d"] += to_float(r.get("Total Impressions"))
        data[team]["Push"]["c"] += to_float(r.get("Total Clicks"))
        data[team]["Push"]["v"] += to_float(r.get("Unique Conversions"))

    for r in read_sheet(gc, WEBENGAGE_SHEET_ID, "Inapp_Campaign"):
        if not is_current_month(r.get("Reporting Period Start Date", "")): continue
        team = map_team_webengage(r.get("Type of Campaign", ""), r.get("Campaign Name", ""))
        if team not in TEAMS: continue
        data[team]["In-app"]["d"] += to_float(r.get("Unique Impressions"))
        data[team]["In-app"]["c"] += to_float(r.get("Unique Clicks"))
        data[team]["In-app"]["v"] += to_float(r.get("Unique Conversions"))

    h_total = ["Team", "Delivered / Impr", "Clicks", "CTR", "Conversions"]
    total_rows = []
    for team in TEAMS:
        td = sum(data[team][ch]["d"] for ch in CHANNELS)
        tc = sum(data[team][ch]["c"] for ch in CHANNELS)
        tv = sum(data[team][ch]["v"] for ch in CHANNELS)
        total_rows.append([team, int(td), int(tc), safe_pct(tc, td), int(tv)])

    h_ch = ["Team", "Channel", "Delivered / Impr", "Clicks", "CTR", "Conversions"]
    ch_rows = []
    for team in TEAMS:
        for ch in CHANNELS:
            d = data[team][ch]
            ch_rows.append([team, ch, int(d["d"]), int(d["c"]), safe_pct(d["c"], d["d"]), int(d["v"])])

    return h_total, total_rows, h_ch, ch_rows


# ══════════════════════════════════════════════════════════════════════════════
# REPORT 4 — Journey Funnel
# ══════════════════════════════════════════════════════════════════════════════

def build_journey_funnel(gc):
    print("[Report 4] Journey Funnel...")
    rows  = read_sheet(gc, WEBENGAGE_SHEET_ID, "Journey_Campaign")
    by_jc = defaultdict(lambda: {"sent": 0, "deliv": 0, "impr": 0, "clicks": 0, "conv": 0})
    by_j  = defaultdict(lambda: {"sent": 0, "deliv": 0, "impr": 0, "clicks": 0, "conv": 0})
    # {(journey, language): metrics} — language from campaign name
    by_jl = defaultdict(lambda: {"sent": 0, "deliv": 0, "impr": 0, "clicks": 0, "conv": 0})

    for r in rows:
        if not is_current_month(r.get("Reporting Period Start Date", "")): continue
        journey  = r.get("Journey Name", "").strip() or r.get("Campaign Name", "").strip()
        channel  = r.get("Channel", "").strip()
        campaign = r.get("Campaign Name", "").strip()
        if not journey: continue
        sent  = to_float(r.get("Sent"))
        deliv = to_float(r.get("Delivered"))
        impr  = to_float(r.get("Unique Impressions"))
        click = to_float(r.get("Unique Clicks"))
        conv  = to_float(r.get("Unique Conversions"))

        by_jc[(journey, channel)]["sent"]   += sent
        by_jc[(journey, channel)]["deliv"]  += deliv
        by_jc[(journey, channel)]["impr"]   += impr
        by_jc[(journey, channel)]["clicks"] += click
        by_jc[(journey, channel)]["conv"]   += conv

        by_j[journey]["sent"]   += sent
        by_j[journey]["deliv"]  += deliv
        by_j[journey]["impr"]   += impr
        by_j[journey]["clicks"] += click
        by_j[journey]["conv"]   += conv

        # Language from campaign name
        lang = get_language(campaign)
        by_jl[(journey, lang)]["sent"]   += sent
        by_jl[(journey, lang)]["deliv"]  += deliv
        by_jl[(journey, lang)]["impr"]   += impr
        by_jl[(journey, lang)]["clicks"] += click
        by_jl[(journey, lang)]["conv"]   += conv

    jc_h = ["Journey Name", "Channel", "Sent", "Delivered / Impr", "Clicks", "CTR", "Conversions"]
    jc_rows = []
    for (journey, channel), d in sorted(by_jc.items()):
        is_inapp   = channel.lower() in ("in-app notification", "in-app")
        deliv_impr = int(d["impr"]) if is_inapp else int(d["deliv"])
        sent_val   = "N/A" if is_inapp else int(d["sent"])
        jc_rows.append([journey, channel, sent_val, deliv_impr,
                        int(d["clicks"]), safe_pct(d["clicks"], deliv_impr), int(d["conv"])])

    j_h = ["Journey Name", "Sent", "Delivered / Impr", "Clicks", "CTR", "Conversions"]
    j_rows = []
    for journey, d in sorted(by_j.items(), key=lambda x: -x[1]["conv"]):
        deliv = int(d["deliv"] or d["impr"])
        j_rows.append([journey, int(d["sent"]), deliv,
                       int(d["clicks"]), safe_pct(d["clicks"], deliv), int(d["conv"])])

    print(f"  {len(by_jc)} journey+channel | {len(by_j)} journeys")
    return jc_h, jc_rows, j_h, j_rows, by_jl


def get_language(name):
    n = name.lower()
    if "hindi" in n:
        return "Hindi"
    if "english" in n:
        return "English"
    return "Other"


def build_classified_journey(by_jl):
    """Aggregate by (journey category, language) using campaign-name language.
    If a journey+language combo is 'Other' AND its Sent > 5000, reclassify as Hindi."""

    # Step 1: Aggregate by (journey, lang) first to check sent per entry
    raw = defaultdict(lambda: {"sent": 0, "deliv": 0, "clicks": 0, "conv": 0})
    for (journey, lang), d in by_jl.items():
        cat = classify_journey(journey)
        raw[(cat, lang)]["sent"]   += d["sent"]
        raw[(cat, lang)]["deliv"]  += d["deliv"] or d["impr"]
        raw[(cat, lang)]["clicks"] += d["clicks"]
        raw[(cat, lang)]["conv"]   += d["conv"]

    # Step 2: Rebuild with reclassification — Other → Hindi if Sent > 5000
    final = defaultdict(lambda: {"sent": 0, "deliv": 0, "clicks": 0, "conv": 0})
    for (cat, lang), d in raw.items():
        effective_lang = "Hindi" if (lang == "Other" and d["sent"] > 5000) else lang
        for metric in ["sent", "deliv", "clicks", "conv"]:
            final[(cat, effective_lang)][metric] += d[metric]

    h = ["Journey Category", "Language", "Sent", "Delivered / Impr", "Clicks", "CTR", "Conversions"]
    rows = []
    for (cat, lang), d in sorted(final.items(), key=lambda x: (-x[1]["conv"], x[0][0], x[0][1])):
        rows.append([cat, lang, int(d["sent"]), int(d["deliv"]),
                     int(d["clicks"]), safe_pct(d["clicks"], d["deliv"]), int(d["conv"])])
    return h, rows


# ══════════════════════════════════════════════════════════════════════════════
# REPORT 5 — WhatsApp Failure Analysis
# ══════════════════════════════════════════════════════════════════════════════

def build_wa_failure(gc):
    print("[Report 5] WA Failure Analysis...")
    rows = read_sheet(gc, WEBENGAGE_SHEET_ID, "Whatsapp_Campaign")
    FAIL_COLS = [
        "Failed (DND Queue Drop)", "Failed (Frequency Capping Queue Drop)",
        "Failed (Personalization Error)", "Failed (Channel Not Available)",
        "Failed (Invalid WhatsApp Number)", "Failed (Message Format Error)",
        "Failed (Template Does Not Exist)", "Failed (Throttling Error)",
        "Failed (User Blocked Messages)", "Failed (User Did Not Initiate Session)",
        "Failed (User Offline)", "Failed (WSP Configuration Error)",
        "Failed (WSP Quote Limit Reached)", "Failed (Time Zone Elapsed)",
        "Failed (Other Failures)",
    ]
    totals       = defaultdict(float)
    total_failed = 0
    for r in rows:
        if not is_current_month(r.get("Reporting Period Start Date", "")): continue
        total_failed += to_float(r.get("Failed"))
        for col in FAIL_COLS:
            totals[col] += to_float(r.get(col))

    h = ["Failure Type", "Count", "% of Total Failed"]
    fail_rows = []
    for col in FAIL_COLS:
        count = int(totals[col])
        label = col.replace("Failed (", "").replace(")", "")
        fail_rows.append([label, count, safe_pct(count, total_failed)])
    fail_rows.sort(key=lambda x: -x[1])
    fail_rows.append(["Total Failed", int(total_failed), "100%"])
    print(f"  Total failed: {int(total_failed)}")
    return h, fail_rows


# ══════════════════════════════════════════════════════════════════════════════
# REPORT 7 — Push Notification Engagement
# ══════════════════════════════════════════════════════════════════════════════

def build_push_engagement(gc):
    print("[Report 7] Push Engagement...")
    rows = read_sheet(gc, WEBENGAGE_SHEET_ID, "Push_Campaign")

    total_impr = total_click = total_dismiss = total_conv = 0
    by_seg  = defaultdict(lambda: {"i": 0, "c": 0, "v": 0})
    by_type = defaultdict(lambda: {"i": 0, "c": 0, "v": 0})

    for r in rows:
        if not is_current_month(r.get("Reporting Period Start Date", "")): continue
        impr    = to_float(r.get("Total Impressions"))
        click   = to_float(r.get("Total Clicks"))
        dismiss = to_float(r.get("Total Dismisses") or r.get("Unique Dismisses"))
        conv    = to_float(r.get("Unique Conversions"))
        segment = r.get("Segment Name", "").strip() or "All Users"

        total_impr    += impr
        total_click   += click
        total_dismiss += dismiss
        total_conv    += conv

        by_seg[segment]["i"] += impr
        by_seg[segment]["c"] += click
        by_seg[segment]["v"] += conv

        # Campaign type — One-time only
        camp_type = r.get("Type of Campaign", "").strip().lower()
        if camp_type in ("one-time", "one time"):
            ctype = extract_push_type(r.get("Campaign Name", ""))
            by_type[ctype]["i"] += impr
            by_type[ctype]["c"] += click
            by_type[ctype]["v"] += conv

    h_ov = ["Metric", "Value"]
    ov_rows = [
        ["Total Impressions",  int(total_impr)],
        ["Total Clicks",       int(total_click)],
        ["CTR",                safe_pct(total_click,   total_impr)],
        ["Total Dismisses",    int(total_dismiss)],
        ["Dismiss Rate",       safe_pct(total_dismiss, total_impr)],
        ["Total Conversions",  int(total_conv)],
        ["Conversion Rate",    safe_pct(total_conv,    total_impr)],
    ]

    h_seg = ["Segment", "Impressions", "Clicks", "CTR", "Conversions"]
    seg_rows = []
    for seg, d in sorted(by_seg.items(), key=lambda x: -x[1]["v"]):
        seg_rows.append([seg, int(d["i"]), int(d["c"]), safe_pct(d["c"], d["i"]), int(d["v"])])

    h_type = ["Campaign Type", "Impressions", "Clicks", "CTR", "Conversions"]
    type_rows = []
    for ctype, d in sorted(by_type.items(), key=lambda x: -x[1]["i"]):
        type_rows.append([ctype, int(d["i"]), int(d["c"]), safe_pct(d["c"], d["i"]), int(d["v"])])

    print(f"  impr={int(total_impr)} CTR={safe_pct(total_click, total_impr)} Dismiss={safe_pct(total_dismiss, total_impr)}")
    return h_ov, ov_rows, h_seg, seg_rows, h_type, type_rows


# ══════════════════════════════════════════════════════════════════════════════
# Assemble rows
# ══════════════════════════════════════════════════════════════════════════════

def section(title, headers, rows):
    return [[title]] + [headers] + rows + [[]]


def assemble_google_rows(now_ist_str, month_label,
                         r1_h, r1_rows,
                         r2_h, r2_total, r2_ch_h, r2_ch_rows,
                         r4_jc_h, r4_jc_rows, r4_j_h, r4_j_rows,
                         r5_h, r5_rows,
                         r7_ov_h, r7_ov_rows, r7_seg_h, r7_seg_rows, r7_type_h, r7_type_rows):
    rows = [[f"Last Updated: {now_ist_str}"], [f"Period: {month_label}"], []]
    rows += section("1. Channel Performance",                          r1_h,     r1_rows)
    rows += section("2. Team Engagement — Total",                      r2_h,     r2_total)
    rows += section("2. Team Engagement — By Channel",                 r2_ch_h,  r2_ch_rows)
    rows += section("4. Journey Funnel — By Journey & Channel",        r4_jc_h,  r4_jc_rows)
    rows += section("4. Journey Funnel — By Journey (Aggregated)",     r4_j_h,   r4_j_rows)
    rows += section("5. WhatsApp Failure Analysis",                    r5_h,     r5_rows)
    rows += section("7. Push Engagement — Overall",                    r7_ov_h,  r7_ov_rows)
    rows += section("7. Push Engagement — By Segment",                 r7_seg_h, r7_seg_rows)
    rows += section("7. Push Engagement — By Campaign Type (One-time)",r7_type_h,r7_type_rows)
    return rows


def assemble_excel_rows(now_ist_str, month_label,
                        r1_h, r1_rows,
                        r2_h, r2_total, r2_ch_h, r2_ch_rows,
                        r4_cl_h, r4_cl_rows,
                        r5_h, r5_rows,
                        r7_ov_h, r7_ov_rows, r7_type_h, r7_type_rows):
    rows = [[f"Last Updated: {now_ist_str}"], [f"Period: {month_label}"], []]
    rows += section("1. Channel Performance",                          r1_h,     r1_rows)
    rows += section("2. Team Engagement — Total",                      r2_h,     r2_total)
    rows += section("2. Team Engagement — By Channel",                 r2_ch_h,  r2_ch_rows)
    rows += section("4. Journey Funnel — Classified",                  r4_cl_h,  r4_cl_rows)
    rows += section("5. WhatsApp Failure Analysis",                    r5_h,     r5_rows)
    rows += section("7. Push Engagement — Overall",                    r7_ov_h,  r7_ov_rows)
    rows += section("7. Push Engagement — By Campaign Type (One-time)",r7_type_h,r7_type_rows)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# Write Google Sheet
# ══════════════════════════════════════════════════════════════════════════════

def write_report(gc, all_rows):
    sh = gc.open_by_key(REPORT_SHEET_ID)
    try:
        ws = sh.worksheet(REPORT_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=REPORT_TAB, rows=2000, cols=15)
    ws.clear()
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print(f"[Sheets] Written {len(all_rows)} rows to '{REPORT_TAB}'")


# ══════════════════════════════════════════════════════════════════════════════
# Write Excel
# ══════════════════════════════════════════════════════════════════════════════

def write_excel(excel_rows):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os as _os

    if not _os.path.exists(ONEDRIVE_PATH):
        print(f"[Excel] File not found: {ONEDRIVE_PATH}")
        return

    wb = load_workbook(ONEDRIVE_PATH)
    if REPORT_TAB in wb.sheetnames:
        del wb[REPORT_TAB]
    ws = wb.create_sheet(REPORT_TAB)

    SECTION_LABELS = {
        "1. Channel Performance",
        "2. Team Engagement — Total",
        "2. Team Engagement — By Channel",
        "4. Journey Funnel — Classified",
        "5. WhatsApp Failure Analysis",
        "7. Push Engagement — Overall",
        "7. Push Engagement — By Campaign Type (One-time)",
    }
    HEADER_FIRST_COLS = {
        "Channel", "Team", "Journey Category", "Failure Type",
        "Metric", "Campaign Type", "Language"
    }

    sec_font  = Font(bold=True, color="FFFFFF")
    sec_fill  = PatternFill("solid", fgColor="2E75B6")
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    center    = Alignment(horizontal="center")

    cr = 1
    for row in excel_rows:
        if not row or not any(str(c).strip() for c in row):
            cr += 1
            continue
        first      = str(row[0]).strip()
        is_section = first in SECTION_LABELS
        is_header  = first in HEADER_FIRST_COLS

        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=cr, column=ci)
            try:
                cell.value = float(str(val).replace('%', '')) if (
                    val not in ("", None, "N/A") and
                    not str(val).endswith('%') and
                    not is_section and not is_header
                ) else val
            except (ValueError, TypeError):
                cell.value = val
            if is_section:
                cell.font = sec_font; cell.fill = sec_fill; cell.alignment = center
            elif is_header:
                cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
        cr += 1

    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 60)

    wb.save(ONEDRIVE_PATH)
    print(f"[Excel] Written '{REPORT_TAB}' to OneDrive")


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 55)
    print("pull_webengage_analysis.py — WebEngage Analysis")
    print("=" * 55)

    gc          = get_gc()
    now_ist     = datetime.now(IST)
    now_ist_str = now_ist.strftime("%d-%m-%Y %H:%M:%S")
    month_label = now_ist.strftime("%B %Y")
    print(f"[Period] {month_label}\n")

    r1_h,  r1_rows                               = build_channel_performance(gc)
    r2_h,  r2_total, r2_ch_h, r2_ch_rows        = build_team_engagement(gc)
    r4_jc_h, r4_jc_rows, r4_j_h, r4_j_rows, r4_by_jl = build_journey_funnel(gc)
    r4_cl_h, r4_cl_rows                               = build_classified_journey(r4_by_jl)
    r5_h,  r5_rows                               = build_wa_failure(gc)
    r7_ov_h, r7_ov_rows, r7_seg_h, r7_seg_rows, r7_type_h, r7_type_rows = build_push_engagement(gc)

    # Google Sheet — all tables
    google_rows = assemble_google_rows(
        now_ist_str, month_label,
        r1_h, r1_rows,
        r2_h, r2_total, r2_ch_h, r2_ch_rows,
        r4_jc_h, r4_jc_rows, r4_j_h, r4_j_rows,
        r5_h, r5_rows,
        r7_ov_h, r7_ov_rows, r7_seg_h, r7_seg_rows, r7_type_h, r7_type_rows
    )
    write_report(gc, google_rows)

    # Excel — classified journey only, no segment table
    excel_rows = assemble_excel_rows(
        now_ist_str, month_label,
        r1_h, r1_rows,
        r2_h, r2_total, r2_ch_h, r2_ch_rows,
        r4_cl_h, r4_cl_rows,
        r5_h, r5_rows,
        r7_ov_h, r7_ov_rows, r7_type_h, r7_type_rows
    )
    write_excel(excel_rows)

    print("\n[OK] pull_webengage_analysis.py complete")


if __name__ == "__main__":
    main()