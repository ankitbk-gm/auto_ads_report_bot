"""
pull_campaign_performance_report.py
Builds a combined Google + Meta campaign performance report with Apptrove event breakdowns.

Reads from:
  - GOOGLE_SHEET_ID: Google_Ads, Meta_Ads, Apptrove_MMP
  - Google Unique Users sheet (GOOGLE_UNIQUE_USERS_SHEET_ID): campaign_unique tab (Google Reach)
  - Meta Graph API: live MTD reach per campaign

Writes to:
  - REPORT_SHEET_ID: Campaign_Performance_Report tab
  - OneDrive Excel: Team Wise Cost.xlsx -> Campaign_Performance_Report sheet

Period: MTD only (single snapshot, no MTD-1/Change — column list didn't ask for those)
Rows: one per (Channel, Campaign) with MTD spend > 0. Same campaign name on both
platforms produces two rows. Zero-spend campaigns (e.g. paused, not yet live this
month) are excluded entirely — not shown as 0 rows.

Objective classification: reuses map_team_google / map_team_meta VERBATIM from
pull_team_cost_report.py (per your choice), stripped of the "Marketing " prefix.
NOTE: map_team_meta can also return "Superstar" or "Supply" for non-marketing Meta
campaigns sharing the ad account. Those rows are excluded here since this report
is Onboarding/Retention only — see build_rows().

Reach: Google from campaign_unique sheet (same source as other reports, summed per
campaign instead of one grand total). Meta via live Graph API MTD query (per your
choice — matches pull_retention_report.py's fetch_meta_mtd_reach methodology).

Google Clicks: NOT CONFIRMED IN CODE. No existing script in this repo reads a
clicks-type column from Google_Ads — every read only touches Campaign, Ad_Group,
Date, Spend_INR, Impressions. See GOOGLE_CLICKS_COLUMN below — if your Google_Ads
tab has clicks under a different name, change the constant and clicks will populate
automatically. Until confirmed, Google rows show "N/A" for Clicks and CTR, and the
script prints a warning at runtime so this isn't silently wrong.
"""

import os
from datetime import datetime, timedelta
from collections import defaultdict
import pytz
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials
import requests as req

load_dotenv()

IST = pytz.timezone("Asia/Kolkata")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

GOOGLE_SHEET_ID         = os.getenv("GOOGLE_SHEET_ID")
REPORT_SHEET_ID         = os.getenv("REPORT_SHEET_ID")
GOOGLE_UNIQUE_USERS_SID = os.getenv("GOOGLE_UNIQUE_USERS_SHEET_ID")
GOOGLE_UNIQUE_USERS_TAB = os.getenv("GOOGLE_UNIQUE_USERS_TAB", "campaign_unique")
ONEDRIVE_PATH = os.getenv("ONEDRIVE_EXCEL_PATH", r"C:\Users\anitb\OneDrive - Agrim Wholesale Private Limited\Marketing_Reports\Team Wise Cost.xlsx")

REPORT_TAB = "Campaign_Performance_Report"

# VERIFY: set this to the actual Google_Ads clicks column name if it has one.
GOOGLE_CLICKS_COLUMN = "Clicks"

# Matches fetch_meta_reach() in pull_retention_report.py / pull_onboarding_report.py /
# pull_kam_brand_report.py. (pull_meta.py's main data pull uses v19.0 — pre-existing
# version inconsistency in the repo, not introduced here.)
META_API_VERSION = "v18.0"

HEADERS = ["Channel", "Campaign Name", "Objective", "Impressions", "Reach", "Clicks", "CTR",
           "App Opened", "View Content", "First Homepage Viewed", "First Purchase Success", "Purchase"]

EMPTY_APP = {"app_opened": 0, "view_content": 0, "first_homePage_viewed": 0,
             "first_purchase_success": 0, "purchase": 0}


# ── Date period ────────────────────────────────────────────────────────────────

def get_mtd():
    """Same AM/PM-aware MTD logic as pull_team_cost_report.py's get_periods()
    (before 12 PM IST: MTD ends yesterday; after: MTD ends today). No MTD-1 needed."""
    now_ist   = datetime.now(IST)
    today     = now_ist.date()
    yesterday = today - timedelta(days=1)
    mtd_start = today.replace(day=1)

    if now_ist.hour < 12:
        mtd_end = yesterday if yesterday.month == today.month else today
    else:
        mtd_end = today

    return mtd_start, mtd_end


def in_period(date_val, start, end):
    """Verbatim copy from pull_team_cost_report.py — handles both date formats
    seen across Google_Ads / Meta_Ads / Apptrove_MMP."""
    if isinstance(date_val, datetime):
        d = date_val.date()
    elif isinstance(date_val, str):
        for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
            try:
                d = datetime.strptime(date_val[:10], fmt).date()
                break
            except ValueError:
                continue
        else:
            return False
    else:
        d = date_val
    return start <= d <= end


# ── Auth ───────────────────────────────────────────────────────────────────────

def get_gc():
    creds_path = (
        os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH")
        or os.getenv("GOOGLE_CREDENTIALS_PATH")
        or "service_account.json"
    )
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


def read_sheet(gc, sheet_id, tab_name):
    sh   = gc.open_by_key(sheet_id)
    ws   = sh.worksheet(tab_name)
    rows = ws.get_all_values()
    if not rows or len(rows) < 2:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


# ── Objective classification (verbatim from pull_team_cost_report.py) ──────────

def map_team_meta(campaign_name):
    n = campaign_name.lower().strip()
    if "superstar" in n or n.startswith("ss_"):                  return "Superstar"
    if "seller" in n:                                            return "Supply"
    if n.startswith("ar_") and "install" in n:                   return "Marketing Onboarding"
    if any(k in n for k in ["onboarding", "aci", "otp_entered", "kyced_but_not_transacted"]): return "Marketing Onboarding"
    if any(k in n for k in ["retention", "ace"]):                return "Marketing Retention"
    if "marketing" in n:                                         return "Marketing Retention"
    return "Marketing Retention"


def map_team_google(campaign_name):
    n = campaign_name.lower().strip()
    if any(k in n for k in ["aci", "otp_entered", "kyced_but_not_transacted", "install", "onboarding"]): return "Marketing Onboarding"
    if any(k in n for k in ["retention", "ace"]):                return "Marketing Retention"
    if "marketing" in n:                                         return "Marketing Retention"
    return "Marketing Retention"


# ── Process Google Ads ─────────────────────────────────────────────────────────

def process_google(gc, mtd_start, mtd_end):
    print("[Google] Reading Google_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Google_Ads")
    clicks_available = bool(rows) and GOOGLE_CLICKS_COLUMN in rows[0]
    if not clicks_available:
        print(f"  WARNING: column '{GOOGLE_CLICKS_COLUMN}' not found in Google_Ads — Clicks/CTR will show N/A for all Google rows")

    data = defaultdict(lambda: {"impressions": 0, "clicks": 0, "spend": 0.0})

    for row in rows:
        campaign = row.get("Campaign", "").strip()
        if not campaign:
            continue
        if not in_period(row.get("Date", ""), mtd_start, mtd_end):
            continue
        try:
            impr = int(float(row.get("Impressions", 0) or 0))
        except ValueError:
            impr = 0
        try:
            spend = float(row.get("Spend_INR", 0) or 0)
        except ValueError:
            spend = 0.0
        data[campaign]["impressions"] += impr
        data[campaign]["spend"]       += spend

        if clicks_available:
            try:
                data[campaign]["clicks"] += int(float(row.get(GOOGLE_CLICKS_COLUMN, 0) or 0))
            except ValueError:
                pass

    print(f"  {len(data)} Google campaigns in MTD window (before spend filter)")
    return data, clicks_available


# ── Process Meta Ads ───────────────────────────────────────────────────────────

def process_meta(gc, mtd_start, mtd_end):
    print("[Meta] Reading Meta_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Meta_Ads")

    data = defaultdict(lambda: {"impressions": 0, "clicks": 0, "spend": 0.0})

    for row in rows:
        campaign = row.get("Campaign", "").strip()
        if not campaign:
            continue
        if not in_period(row.get("Date", ""), mtd_start, mtd_end):
            continue
        try:
            impr = int(float(row.get("Impressions", 0) or 0))
        except ValueError:
            impr = 0
        try:
            clicks = int(float(row.get("Unique_Outbound_Clicks", 0) or 0))
        except ValueError:
            clicks = 0
        try:
            spend = float(row.get("Spend_INR", 0) or 0)
        except ValueError:
            spend = 0.0
        data[campaign]["impressions"] += impr
        data[campaign]["clicks"]      += clicks
        data[campaign]["spend"]       += spend

    print(f"  {len(data)} Meta campaigns in MTD window (before spend filter)")
    return data


# ── Reach ──────────────────────────────────────────────────────────────────────

def get_google_reach(gc):
    """Campaign-level reach from the Google Unique Users sheet — same tab
    pull_retention_report.py reads, summed per campaign instead of one grand total."""
    print("[Google] Reading Unique Users sheet...")
    rows = read_sheet(gc, GOOGLE_UNIQUE_USERS_SID, GOOGLE_UNIQUE_USERS_TAB)
    reach = defaultdict(int)
    for row in rows:
        campaign = row.get("Campaign", "").strip()
        uu       = row.get("Unique users", "").strip()
        if not campaign or not uu or uu == "--":
            continue
        try:
            reach[campaign] += int(str(uu).replace(",", ""))
        except ValueError:
            continue
    print(f"  Reach found for {len(reach)} Google campaigns")
    return reach


def get_meta_reach(mtd_start, mtd_end):
    """Live Meta Graph API call for the MTD range — matches
    pull_retention_report.py's fetch_meta_mtd_reach(), per your choice."""
    print("[Meta] Fetching reach from Graph API...")
    access_token = os.getenv("META_ACCESS_TOKEN")
    ad_account   = os.getenv("META_AD_ACCOUNT_ID")
    reach = {}
    if not access_token or not ad_account:
        print("  WARNING: META_ACCESS_TOKEN / META_AD_ACCOUNT_ID missing — Meta reach will be 0")
        return reach
    url    = f"https://graph.facebook.com/{META_API_VERSION}/{ad_account}/insights"
    params = {
        "access_token": access_token,
        "level":        "campaign",
        "fields":       "campaign_name,reach",
        "time_range":   f'{{"since":"{mtd_start}","until":"{mtd_end}"}}',
        "limit":        500,
    }
    try:
        data = req.get(url, params=params, timeout=30).json().get("data", [])
        for item in data:
            campaign = item.get("campaign_name", "").strip()
            reach[campaign] = reach.get(campaign, 0) + int(item.get("reach", 0) or 0)
    except Exception as e:
        print(f"  [Meta Reach] Error: {e}")
    print(f"  Reach found for {len(reach)} Meta campaigns")
    return reach


# ── Process Apptrove ───────────────────────────────────────────────────────────

def process_apptrove(gc, mtd_start, mtd_end):
    print("[Apptrove] Reading Apptrove_MMP...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Apptrove_MMP")

    def empty():
        return {"app_opened": 0, "view_content": 0, "first_homePage_viewed": 0,
                 "first_purchase_success": 0, "purchase": 0}

    google_data = defaultdict(empty)
    meta_data   = defaultdict(empty)

    for row in rows:
        partner  = row.get("partner", "").strip()
        channel  = row.get("channel", "").strip()
        campaign = row.get("campaign", "").strip()
        if not campaign or not in_period(row.get("Date", ""), mtd_start, mtd_end):
            continue

        if partner == "Google Ads (Adwords)":
            if not channel or channel == "-":
                continue
            target = google_data[campaign]
        elif partner == "Facebook":
            target = meta_data[campaign]
        else:
            continue  # WhatsApp / others not in scope for this report

        for col in target:
            try:
                target[col] += int(row.get(col, 0) or 0)
            except (ValueError, TypeError):
                pass

    print(f"  Apptrove matched to {len(google_data)} Google + {len(meta_data)} Meta campaigns")
    return google_data, meta_data


# ── Build rows ─────────────────────────────────────────────────────────────────

def safe_ctr(clicks, impressions):
    if clicks == "N/A" or not impressions:
        return "N/A"
    return round(clicks / impressions * 100, 2)


def build_rows(google_data, clicks_available, meta_data, google_reach, meta_reach,
               google_app, meta_app):
    rows = []
    google_skipped_spend = 0

    for campaign, d in sorted(google_data.items(), key=lambda x: -x[1]["impressions"]):
        if d["spend"] <= 0:
            google_skipped_spend += 1
            continue
        objective = map_team_google(campaign).replace("Marketing ", "")
        clicks    = d["clicks"] if clicks_available else "N/A"
        app       = google_app.get(campaign, EMPTY_APP)
        rows.append([
            "Google", campaign, objective, d["impressions"], google_reach.get(campaign, 0),
            clicks, safe_ctr(clicks, d["impressions"]),
            app["app_opened"], app["view_content"], app["first_homePage_viewed"],
            app["first_purchase_success"], app["purchase"],
        ])

    meta_skipped_spend     = 0
    meta_skipped_objective = 0
    for campaign, d in sorted(meta_data.items(), key=lambda x: -x[1]["impressions"]):
        if d["spend"] <= 0:
            meta_skipped_spend += 1
            continue
        objective = map_team_meta(campaign)
        if objective not in ("Marketing Onboarding", "Marketing Retention"):
            meta_skipped_objective += 1
            continue  # Superstar / Supply campaigns sharing the Meta account — out of scope
        objective = objective.replace("Marketing ", "")
        app       = meta_app.get(campaign, EMPTY_APP)
        rows.append([
            "Meta", campaign, objective, d["impressions"], meta_reach.get(campaign, 0),
            d["clicks"], safe_ctr(d["clicks"], d["impressions"]),
            app["app_opened"], app["view_content"], app["first_homePage_viewed"],
            app["first_purchase_success"], app["purchase"],
        ])

    if google_skipped_spend:
        print(f"  [Google] Skipped {google_skipped_spend} campaigns with zero MTD spend")
    if meta_skipped_spend:
        print(f"  [Meta] Skipped {meta_skipped_spend} campaigns with zero MTD spend")
    if meta_skipped_objective:
        print(f"  [Meta] Skipped {meta_skipped_objective} non-marketing campaigns (Superstar/Supply)")

    return rows


# ── Write Google Sheet ─────────────────────────────────────────────────────────

def write_report(gc, data_rows, now_ist_str):
    sh = gc.open_by_key(REPORT_SHEET_ID)
    try:
        ws = sh.worksheet(REPORT_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=REPORT_TAB, rows=500, cols=15)

    all_rows = [[f"Last Updated: {now_ist_str}"], []]
    all_rows.append(HEADERS)
    all_rows.extend(data_rows)

    ws.clear()
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print(f"[Sheets] Written {len(data_rows)} rows to '{REPORT_TAB}'")


# ── Write OneDrive Excel ───────────────────────────────────────────────────────

def write_excel(data_rows, now_ist_str):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os as _os

    if not _os.path.exists(ONEDRIVE_PATH):
        print(f"[Excel] File not found: {ONEDRIVE_PATH}")
        return

    wb = load_workbook(ONEDRIVE_PATH)
    if REPORT_TAB in wb.sheetnames:
        del wb[REPORT_TAB]
    ws = wb.create_sheet(REPORT_TAB)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center")

    cr = 1
    ws.cell(row=cr, column=1, value=f"Last Updated: {now_ist_str}").font = Font(italic=True)
    cr += 2

    for ci, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=cr, column=ci, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
    cr += 1

    for row in data_rows:
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=cr, column=ci)
            try:
                cell.value = float(val) if val not in ("", None, "N/A") else val
            except (ValueError, TypeError):
                cell.value = val
        cr += 1

    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = ml + 4

    wb.save(ONEDRIVE_PATH)
    print(f"[Excel] Written '{REPORT_TAB}' to OneDrive")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("pull_campaign_performance_report.py — Campaign Performance Report")
    print("=" * 55)

    mtd_start, mtd_end = get_mtd()
    print(f"[Period] MTD: {mtd_start} → {mtd_end}")

    gc          = get_gc()
    now_ist_str = datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S")

    google_data, clicks_available = process_google(gc, mtd_start, mtd_end)
    meta_data                     = process_meta(gc, mtd_start, mtd_end)
    google_app, meta_app          = process_apptrove(gc, mtd_start, mtd_end)
    google_reach                  = get_google_reach(gc)
    meta_reach                    = get_meta_reach(str(mtd_start), str(mtd_end))

    data_rows = build_rows(google_data, clicks_available, meta_data,
                            google_reach, meta_reach, google_app, meta_app)

    write_report(gc, data_rows, now_ist_str)
    write_excel(data_rows, now_ist_str)

    print(f"\n✅ pull_campaign_performance_report.py complete — {len(data_rows)} rows")


if __name__ == "__main__":
    main()