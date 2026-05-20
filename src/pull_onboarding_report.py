"""
pull_onboarding_report.py
Builds the Onboarding Report with KYC and First Txn tables.
Three tables: MTD | MTD-1 | Change (absolute)

Timing:
  Before 12 PM IST: MTD = May 1 → yesterday,  MTD-1 = May 1 → day before yesterday
  After  12 PM IST: MTD = May 1 → today,       MTD-1 = May 1 → yesterday
"""

import os
import requests as req
from datetime import datetime, timedelta
from collections import defaultdict
import pytz
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials

load_dotenv()

IST = pytz.timezone("Asia/Kolkata")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

GOOGLE_SHEET_ID         = os.getenv("GOOGLE_SHEET_ID")
WEBENGAGE_SHEET_ID      = os.getenv("WEBENGAGE_SHEET_ID")
REPORT_SHEET_ID         = os.getenv("REPORT_SHEET_ID")
GOOGLE_UNIQUE_USERS_SID = os.getenv("GOOGLE_UNIQUE_USERS_SHEET_ID")
GOOGLE_UNIQUE_USERS_TAB = os.getenv("GOOGLE_UNIQUE_USERS_TAB", "campaign_unique")
ONEDRIVE_PATH           = r"C:\Users\anitb\OneDrive - Agrim Wholesale Private Limited\Marketing_Reports\Team Wise Cost.xlsx"

WATI_COST_RATE      = 0.96
WEBENGAGE_BASE_RATE = 0.87
WEBENGAGE_GST       = 1.18
GOOGLE_META_GST     = 1.18

WATI_KYC_KEYWORDS = ["otp_entered_not_kyc", "marketing_install"]


# ── Auth ───────────────────────────────────────────────────────────────────────

def get_gc():
    creds_path = (
        os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH")
        or os.getenv("GOOGLE_CREDENTIALS_PATH")
        or "service_account.json"
    )
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


# ── Periods ────────────────────────────────────────────────────────────────────

def get_periods():
    now_ist   = datetime.now(IST)
    today     = now_ist.date()
    yesterday = today - timedelta(days=1)
    mtd_start = today.replace(day=1)
    if now_ist.hour < 12:
        mtd_end  = yesterday
        mtd1_end = yesterday - timedelta(days=1)
    else:
        mtd_end  = today
        mtd1_end = yesterday
    return mtd_start, mtd_end, mtd1_end


def in_range(date_val, start, end):
    if isinstance(date_val, str):
        for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
            try:
                d = datetime.strptime(date_val[:10], fmt).date()
                return start <= d <= end
            except ValueError:
                continue
        return False
    if hasattr(date_val, 'date'):
        return start <= date_val.date() <= end
    return start <= date_val <= end


# ── Read sheet ─────────────────────────────────────────────────────────────────

def read_sheet(gc, sheet_id, tab_name):
    sh   = gc.open_by_key(sheet_id)
    ws   = sh.worksheet(tab_name)
    rows = ws.get_all_values()
    if not rows or len(rows) < 2:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


# ── Campaign classification ────────────────────────────────────────────────────

def is_google_kyc(c): n = c.lower(); return any(k in n for k in ["otp_entered", "aci"])
def is_google_txn(c): return "kyced_but_not_transacted" in c.lower()
def is_meta_kyc(c):   n = c.lower(); return any(k in n for k in ["install","otp_entered"]) and n.startswith("ar_onboarding")
def is_meta_txn(c):   n = c.lower(); return "kyc" in n and n.startswith("ar_onboarding")
def is_wati_kyc(c):   n = c.lower(); return any(k in n for k in WATI_KYC_KEYWORDS)
def is_wati_ob(c):    return "marketing_onboarding" in c.lower()
def is_we_kyc(j):     return "kyc_journey" in j.lower()
def is_we_ob(t, n):
    t = t.lower().strip(); n = n.lower().strip()
    if t == "relay":    return False
    if t == "journey":  return "add to cart but not purchased" not in n
    if t == "one time": return "onboarding" in n
    return False


# ── Process Google ─────────────────────────────────────────────────────────────

def process_google(gc, mtd_start, mtd_end, mtd1_end):
    print("[Google] Reading Google_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Google_Ads")

    kyc_s=0.; txn_s=0.; kyc_s1=0.; txn_s1=0.
    kyc_camps = set(); txn_camps = set()

    for row in rows:
        c = row.get("Campaign","").strip()
        d = row.get("Date","")
        try: spend = float(row.get("Spend_INR",0) or 0) * GOOGLE_META_GST
        except: spend = 0
        if spend <= 0: continue

        if is_google_kyc(c):
            if in_range(d, mtd_start, mtd_end):  kyc_s += spend; kyc_camps.add(c)
            if in_range(d, mtd_start, mtd1_end): kyc_s1 += spend
        elif is_google_txn(c):
            if in_range(d, mtd_start, mtd_end):  txn_s += spend; txn_camps.add(c)
            if in_range(d, mtd_start, mtd1_end): txn_s1 += spend

    print(f"  KYC spend MTD:{round(kyc_s,2)} MTD-1:{round(kyc_s1,2)} | TXN MTD:{round(txn_s,2)} MTD-1:{round(txn_s1,2)}")
    return {"kyc_spend":round(kyc_s,2), "kyc_spend1":round(kyc_s1,2),
            "txn_spend":round(txn_s,2), "txn_spend1":round(txn_s1,2),
            "kyc_camps":kyc_camps, "txn_camps":txn_camps}


def process_google_unique_users(gc, kyc_camps, txn_camps):
    print("[Google] Reading Unique Users...")
    rows    = read_sheet(gc, GOOGLE_UNIQUE_USERS_SID, GOOGLE_UNIQUE_USERS_TAB)
    kyc_r=0; txn_r=0
    for row in rows:
        c  = row.get("Campaign","").strip()
        uu = row.get("Unique users","").strip()
        if not uu or uu == "--": continue
        try: uu = int(str(uu).replace(",",""))
        except: continue
        if c in kyc_camps:   kyc_r += uu
        elif c in txn_camps: txn_r += uu
    txn_r = min(txn_r, 78990)
    print(f"  KYC reach:{kyc_r} | TXN reach:{txn_r}")
    return kyc_r, txn_r


# ── Process Meta ───────────────────────────────────────────────────────────────

def process_meta(gc, mtd_start, mtd_end, mtd1_end):
    print("[Meta] Reading Meta_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Meta_Ads")

    kyc_s=0.; txn_s=0.; kyc_s1=0.; txn_s1=0.
    kyc_camps = set(); txn_camps = set()

    for row in rows:
        c = row.get("Campaign","").strip()
        d = row.get("Date","")
        try: spend = float(row.get("Spend_INR",0) or 0) * GOOGLE_META_GST
        except: spend = 0
        if spend <= 0: continue

        if is_meta_kyc(c):
            if in_range(d, mtd_start, mtd_end):  kyc_s += spend; kyc_camps.add(c)
            if in_range(d, mtd_start, mtd1_end): kyc_s1 += spend
        elif is_meta_txn(c):
            if in_range(d, mtd_start, mtd_end):  txn_s += spend; txn_camps.add(c)
            if in_range(d, mtd_start, mtd1_end): txn_s1 += spend

    return {"kyc_spend":round(kyc_s,2), "kyc_spend1":round(kyc_s1,2),
            "txn_spend":round(txn_s,2), "txn_spend1":round(txn_s1,2),
            "kyc_camps":kyc_camps, "txn_camps":txn_camps}


def fetch_meta_reach(mtd_start, mtd_end, kyc_camps, txn_camps):
    access_token = os.getenv("META_ACCESS_TOKEN")
    ad_account   = os.getenv("META_AD_ACCOUNT_ID")
    if not access_token or not ad_account:
        return 0, 0
    url    = f"https://graph.facebook.com/v18.0/{ad_account}/insights"
    params = {"access_token": access_token, "level": "campaign",
              "fields": "campaign_name,reach",
              "time_range": f'{{"since":"{mtd_start}","until":"{mtd_end}"}}',
              "limit": 500}
    kyc_r=0; txn_r=0
    try:
        data = req.get(url, params=params, timeout=30).json().get("data",[])
        for item in data:
            c = item.get("campaign_name","").strip()
            r = int(item.get("reach",0) or 0)
            if c in kyc_camps:   kyc_r += r
            elif c in txn_camps: txn_r += r
    except Exception as e:
        print(f"[Meta Reach] Error: {e}")
    print(f"[Meta Reach] KYC:{kyc_r} TXN:{txn_r}")
    return kyc_r, txn_r


# ── Process Apptrove ───────────────────────────────────────────────────────────

def process_apptrove(gc, mtd_start, mtd_end, mtd1_end, google_kyc, google_txn, meta_txn):
    print("[Apptrove] Reading Apptrove_MMP...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Apptrove_MMP")

    def empty(): return {"google_kyc_fhpv":0,"google_txn_fps":0,"meta_kyc_fhpv":0,"meta_txn_fps":0,"wa_kyc_fhpv":0,"wa_txn_fps":0}
    mtd_r = empty(); mtd1_r = empty()

    for row in rows:
        partner  = row.get("partner","").strip()
        channel  = row.get("channel","").strip()
        campaign = row.get("campaign","").strip()
        date_val = row.get("Date","")
        in_m  = in_range(date_val, mtd_start, mtd_end)
        in_m1 = in_range(date_val, mtd_start, mtd1_end)
        if not in_m and not in_m1: continue

        try:
            fhpv = int(row.get("first_homePage_viewed",0) or 0)
            fps  = int(row.get("first_purchase_success",0) or 0)
        except: fhpv=fps=0

        p = partner.lower(); c = campaign.lower()

        def acc(res):
            if partner == "Google Ads (Adwords)":
                if not channel or channel.strip()=="-": return
                if campaign in google_kyc: res["google_kyc_fhpv"] += fhpv
                elif campaign in google_txn: res["google_txn_fps"] += fps
            elif partner == "Facebook":
                if any(k in c for k in ["install","otp"]): res["meta_kyc_fhpv"] += fhpv
                elif campaign in meta_txn: res["meta_txn_fps"] += fps
            elif "wa" in p or "whatsapp" in p:
                if "voice" in c: return
                res["wa_kyc_fhpv"] += fhpv; res["wa_txn_fps"] += fps

        if in_m:  acc(mtd_r)
        if in_m1: acc(mtd1_r)

    print(f"  MTD  Google KYC FHPV:{mtd_r['google_kyc_fhpv']} TXN FPS:{mtd_r['google_txn_fps']}")
    print(f"  MTD  Meta   KYC FHPV:{mtd_r['meta_kyc_fhpv']}   TXN FPS:{mtd_r['meta_txn_fps']}")
    return mtd_r, mtd1_r


# ── Process WhatsApp ───────────────────────────────────────────────────────────

def process_whatsapp(gc, mtd_start, mtd_end, mtd1_end):
    print("[WhatsApp] Reading Wati and WebEngage...")
    wati_rows = read_sheet(gc, GOOGLE_SHEET_ID, "Wati")

    wk_s=0; wk_c=0.; wt_s=0; wt_c=0.
    wk_s1=0; wk_c1=0.; wt_s1=0; wt_c1=0.

    for row in wati_rows:
        c = row.get("Campaign_Name","").strip()
        d = row.get("Date","")
        try: sent=int(row.get("Sent",0) or 0); dlvd=int(row.get("Delivered",0) or 0)
        except: continue
        cost = dlvd * WATI_COST_RATE

        if is_wati_kyc(c):
            if in_range(d, mtd_start, mtd_end):  wk_s+=sent; wk_c+=cost
            if in_range(d, mtd_start, mtd1_end): wk_s1+=sent; wk_c1+=cost
        elif is_wati_ob(c):
            if in_range(d, mtd_start, mtd_end):  wt_s+=sent; wt_c+=cost
            if in_range(d, mtd_start, mtd1_end): wt_s1+=sent; wt_c1+=cost

    # WebEngage (MTD aggregate — no MTD-1 split possible)
    we_rows = read_sheet(gc, WEBENGAGE_SHEET_ID, "Whatsapp_Campaign")
    now_ist = datetime.now(IST)
    cm = now_ist.month; cy = now_ist.year
    ek_s=0; ek_c=0.; et_s=0; et_c=0.

    for row in we_rows:
        cn = row.get("Campaign Name","").strip()
        ct = row.get("Type of Campaign","").strip()
        jn = row.get("Journey Name","").strip()
        sd = row.get("Reporting Period Start Date","").strip()
        sd_date = None
        for fmt in ("%m/%d/%Y","%Y-%m-%d"):
            try: sd_date=datetime.strptime(sd[:10],fmt); break
            except: continue
        if not sd_date or sd_date.month!=cm or sd_date.year!=cy: continue
        try: sent=int(row.get("Sent",0) or 0); dlvd=int(row.get("Delivered",0) or 0)
        except: continue
        cost = dlvd * WEBENGAGE_BASE_RATE * WEBENGAGE_GST
        if is_we_kyc(jn):       ek_s+=sent; ek_c+=cost
        elif is_we_ob(ct,cn) and not is_we_kyc(jn): et_s+=sent; et_c+=cost

    kyc_sent=wk_s+ek_s; kyc_cost=round(wk_c+ek_c,2)
    txn_sent=wt_s+et_s; txn_cost=round(wt_c+et_c,2)
    kyc_sent1=wk_s1; kyc_cost1=round(wk_c1,2)  # WebEngage excluded from MTD-1
    txn_sent1=wt_s1; txn_cost1=round(wt_c1,2)

    print(f"  KYC MTD sent:{kyc_sent} cost:{kyc_cost} | MTD-1 sent:{kyc_sent1} cost:{kyc_cost1}")
    print(f"  TXN MTD sent:{txn_sent} cost:{txn_cost} | MTD-1 sent:{txn_sent1} cost:{txn_cost1}")
    return kyc_sent,kyc_cost,txn_sent,txn_cost,kyc_sent1,kyc_cost1,txn_sent1,txn_cost1


# ── Build tables ───────────────────────────────────────────────────────────────

def sd(n,d): return "N/A" if not d else round(n/d,2)

def make_kyc_txn_tables(label, g_kyc_r, g_txn_r, g_kyc_s, g_txn_s, app,
                         m_kyc_r, m_kyc_s, m_txn_s,
                         wa_kyc_s_val, wa_kyc_c, wa_txn_s_val, wa_txn_c):
    kh = ["Channel","Target Audience","First Home Page View","Amount Spent","Cost Per KYC"]
    th = ["Channel","Target Audience","First Purchase Success","Amount Spent","Cost Per TXN"]
    kyc = [[f"KYC — {label}"], kh,
           ["Google",   g_kyc_r,    app["google_kyc_fhpv"], g_kyc_s, sd(g_kyc_s, app["google_kyc_fhpv"])],
           ["Meta",     m_kyc_r,    app["meta_kyc_fhpv"],   m_kyc_s, sd(m_kyc_s, app["meta_kyc_fhpv"])],
           ["WhatsApp", wa_kyc_s_val, app["wa_kyc_fhpv"],   wa_kyc_c, sd(wa_kyc_c, app["wa_kyc_fhpv"])]]
    txn = [[f"First Txn — {label}"], th,
           ["Google",   g_txn_r,    app["google_txn_fps"],  g_txn_s, sd(g_txn_s, app["google_txn_fps"])],
           ["Meta",     m_kyc_r,    app["meta_txn_fps"],    m_txn_s, sd(m_txn_s, app["meta_txn_fps"])],
           ["WhatsApp", wa_txn_s_val, app["wa_txn_fps"],    wa_txn_c, sd(wa_txn_c, app["wa_txn_fps"])]]
    return kyc, txn


def make_change_tables(kyc_mtd, txn_mtd, kyc_mtd1, txn_mtd1):
    def chg(t_mtd, t_mtd1):
        rows = []
        for r, r1 in zip(t_mtd, t_mtd1):
            if not r: rows.append(r); continue
            if r[0].startswith("KYC") or r[0].startswith("First Txn"):
                rows.append([r[0].split("—")[0].strip() + " — Change (MTD - MTD-1)"])
                continue
            if r[0] == "Channel": rows.append(r); continue
            cr = [r[0]]
            for i in range(1, len(r)):
                try: cr.append(round(float(str(r[i]).replace("N/A","0") or 0) - float(str(r1[i]).replace("N/A","0") or 0), 2))
                except: cr.append("N/A")
            rows.append(cr)
        return rows
    return chg(kyc_mtd, kyc_mtd1), chg(txn_mtd, txn_mtd1)


# ── Write Google Sheet ─────────────────────────────────────────────────────────

def write_report(gc, tables, now_ist_str, mtd_start, mtd_end, mtd1_end):
    sh = gc.open_by_key(REPORT_SHEET_ID)
    try:
        ws = sh.worksheet("Onboarding_report")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title="Onboarding_report", rows=100, cols=10)

    all_rows = [
        [f"Last Updated: {now_ist_str}"],
        [f"MTD: {mtd_start} → {mtd_end} | MTD-1: {mtd_start} → {mtd1_end}"],
        ["Note: Event numbers are from Apptrove"],
        ["Note: Google Unique Users refreshes at 7 PM IST daily"],
        [],
    ]
    for kyc_t, txn_t in tables:
        all_rows.extend(kyc_t)
        all_rows.append([])
        all_rows.extend(txn_t)
        all_rows.append([])

    ws.clear()
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print(f"[Sheets] Written Onboarding_report tab")


# ── Write OneDrive Excel ───────────────────────────────────────────────────────

def write_excel_onboarding(tables, now_ist_str, mtd_start, mtd_end, mtd1_end):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os as _os

    if not _os.path.exists(ONEDRIVE_PATH):
        print(f"[Excel] File not found: {ONEDRIVE_PATH}")
        return

    wb = load_workbook(ONEDRIVE_PATH)
    if "Onboarding_report" in wb.sheetnames:
        del wb["Onboarding_report"]
    ws = wb.create_sheet("Onboarding_report")

    label_fill  = PatternFill("solid", fgColor="2E75B6")
    label_font  = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    center      = Alignment(horizontal="center")

    cr = 1
    ws.cell(row=cr,column=1,value=f"Last Updated: {now_ist_str}").font = Font(italic=True); cr+=1
    ws.cell(row=cr,column=1,value=f"MTD: {mtd_start} → {mtd_end} | MTD-1: {mtd_start} → {mtd1_end}").font = Font(italic=True); cr+=1
    ws.cell(row=cr,column=1,value="Note: Event numbers are from Apptrove").font = Font(italic=True); cr+=2

    for kyc_t, txn_t in tables:
        for table in [kyc_t, txn_t]:
            for ri, row in enumerate(table):
                for ci, val in enumerate(row, start=1):
                    cell = ws.cell(row=cr, column=ci)
                    try: cell.value = float(val) if val not in ("","N/A",None) else val
                    except: cell.value = val
                    if ri == 0:
                        cell.font = label_font; cell.fill = label_fill; cell.alignment = center
                    elif ri == 1:
                        cell.font = header_font; cell.fill = header_fill; cell.alignment = center
                cr += 1
            cr += 1

    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = ml + 4

    wb.save(ONEDRIVE_PATH)
    print(f"[Excel] Written Onboarding_report to OneDrive")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("pull_onboarding_report.py — Onboarding Report")
    print("=" * 55)

    mtd_start, mtd_end, mtd1_end = get_periods()
    print(f"[Period] MTD: {mtd_start} → {mtd_end} | MTD-1: {mtd_start} → {mtd1_end}")

    gc          = get_gc()
    now_ist_str = datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S")

    google       = process_google(gc, mtd_start, mtd_end, mtd1_end)
    g_kyc_r, g_txn_r = process_google_unique_users(gc, google["kyc_camps"], google["txn_camps"])

    meta         = process_meta(gc, mtd_start, mtd_end, mtd1_end)
    m_kyc_r, _   = fetch_meta_reach(
        mtd_start.strftime("%Y-%m-%d"), mtd_end.strftime("%Y-%m-%d"),
        meta["kyc_camps"], meta["txn_camps"]
    )
    m_kyc_r1, _  = fetch_meta_reach(
        mtd_start.strftime("%Y-%m-%d"), mtd1_end.strftime("%Y-%m-%d"),
        meta["kyc_camps"], meta["txn_camps"]
    )

    app_mtd, app_mtd1 = process_apptrove(
        gc, mtd_start, mtd_end, mtd1_end,
        google["kyc_camps"], google["txn_camps"], meta["txn_camps"]
    )

    wa = process_whatsapp(gc, mtd_start, mtd_end, mtd1_end)
    kyc_sent,kyc_cost,txn_sent,txn_cost,kyc_sent1,kyc_cost1,txn_sent1,txn_cost1 = wa

    mtd_label  = f"MTD ({mtd_start} → {mtd_end})"
    mtd1_label = f"MTD-1 ({mtd_start} → {mtd1_end})"

    kyc_mtd, txn_mtd = make_kyc_txn_tables(
        mtd_label, g_kyc_r, g_txn_r, google["kyc_spend"], google["txn_spend"], app_mtd,
        m_kyc_r, meta["kyc_spend"], meta["txn_spend"],
        kyc_sent, kyc_cost, txn_sent, txn_cost
    )
    kyc_mtd1, txn_mtd1 = make_kyc_txn_tables(
        mtd1_label, g_kyc_r, g_txn_r, google["kyc_spend1"], google["txn_spend1"], app_mtd1,
        m_kyc_r1, meta["kyc_spend1"], meta["txn_spend1"],
        kyc_sent1, kyc_cost1, txn_sent1, txn_cost1
    )
    kyc_chg, txn_chg = make_change_tables(kyc_mtd, txn_mtd, kyc_mtd1, txn_mtd1)

    tables = [(kyc_mtd, txn_mtd), (kyc_mtd1, txn_mtd1), (kyc_chg, txn_chg)]

    write_report(gc, tables, now_ist_str, mtd_start, mtd_end, mtd1_end)
    write_excel_onboarding(tables, now_ist_str, mtd_start, mtd_end, mtd1_end)

    print("\n[OK] pull_onboarding_report.py complete")


if __name__ == "__main__":
    main()