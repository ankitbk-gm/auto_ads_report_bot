"""
pull_kam_brand_report.py
Builds the KAM Brand Performance Report combining Google, Meta and Apptrove data.

Reads from:
  - Original sheet (GOOGLE_SHEET_ID): Google_Ads, Meta_Ads, Apptrove_MMP
  - Google Unique Users sheet (GOOGLE_UNIQUE_USERS_SHEET_ID): campaign_unique tab

Writes to:
  - Report sheet (REPORT_SHEET_ID): KAM_Brand_Report tab
  - OneDrive Excel: Team Wise Cost.xlsx -> KAM_Brand_Report sheet

Period: MTD
Rows: one per brand (dynamic, based on active spend)
PPV columns: blank for now (Mixpanel pending)
"""

import os
import re
from datetime import datetime, timedelta
from collections import defaultdict
import pytz
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials
import requests as req

load_dotenv()

IST = pytz.timezone("Asia/Kolkata")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

GOOGLE_SHEET_ID         = os.getenv("GOOGLE_SHEET_ID")
REPORT_SHEET_ID         = os.getenv("REPORT_SHEET_ID")
GOOGLE_UNIQUE_USERS_SID = os.getenv("GOOGLE_UNIQUE_USERS_SHEET_ID")
GOOGLE_UNIQUE_USERS_TAB = os.getenv("GOOGLE_UNIQUE_USERS_TAB", "campaign_unique")
ONEDRIVE_PATH = os.getenv("ONEDRIVE_EXCEL_PATH", r"C:\Users\anitb\OneDrive - Agrim Wholesale Private Limited\Marketing_Reports\Team Wise Cost.xlsx")

GOOGLE_META_GST = 1.18

# Retention campaign keywords
GOOGLE_RETENTION_KEYWORDS = ["ace", "ar_purchasers"]
META_RETENTION_KEYWORDS   = ["retention"]

HEADERS = ["Brand", "Launch Date", "Paused Date", "Reach", "Impressions", "CTR (App/Impr%)",
           "Spend", "App Traffic", "Avg PPV Before", "Avg PPV After", "Difference"]


# ── Auth ───────────────────────────────────────────────────────────────────────

def get_gc():
    creds_path = (
        os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH")
        or os.getenv("GOOGLE_CREDENTIALS_PATH")
        or "service_account.json"
    )
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


# ── Date period ────────────────────────────────────────────────────────────────

def get_mtd():
    now_ist   = datetime.now(IST)
    today     = now_ist.date()
    yesterday = today - timedelta(days=1)
    mtd_start = today.replace(day=1)
    mtd_end   = today if now_ist.hour >= 12 else yesterday
    return mtd_start, mtd_end


def in_range(date_val, start, end):
    if isinstance(date_val, str):
        for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
            try:
                d = datetime.strptime(date_val[:10], fmt).date()
                return start <= d <= end
            except ValueError:
                continue
        return False
    if hasattr(date_val, 'date'):
        return start <= date_val.date() <= end
    return start <= date_val <= end


# ── Read sheet ─────────────────────────────────────────────────────────────────

def read_sheet(gc, sheet_id, tab_name):
    sh   = gc.open_by_key(sheet_id)
    ws   = sh.worksheet(tab_name)
    rows = ws.get_all_values()
    if not rows or len(rows) < 2:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


# ── Brand extraction ───────────────────────────────────────────────────────────

# Date suffix pattern: _7Apr26, _01May26, _12May26, _20Apr26 etc.
DATE_SUFFIX = re.compile(r'_\d{1,2}[A-Za-z]{3,}\d{2,4}\s*(-\s*Copy\s*\d*)?$', re.IGNORECASE)
# Also handle "– Copy" variants
DATE_SUFFIX2 = re.compile(r'\s*[–-]\s*Copy\s*\d*$', re.IGNORECASE)


def extract_google_brand(adgroup):
    """Extract brand name from Google ad group name."""
    n = adgroup.strip()
    m = re.match(r'[Bb]rand_[Pp]romotion_(.+)', n)
    if not m:
        return None
    brand = m.group(1).strip().replace('_', ' ')
    if re.match(r'^\d+[A-Za-z]+\d+$', brand):
        return None
    return brand


def extract_meta_brand(ad):
    """Extract brand name from Meta ad name."""
    n = ad.strip()
    # Skip subcat ads
    nl = n.lower()
    if any(k in nl for k in ["catalogue", "gibberellic", "plant growth"]):
        return None
    # Remove "– Copy" or "- Copy" suffix first
    n = DATE_SUFFIX2.sub('', n).strip()
    # Remove date suffix
    n = DATE_SUFFIX.sub('', n).strip()
    if not n:
        return None
    # Replace underscores with spaces
    n = n.replace('_', ' ').strip()
    return n


def is_retention_google(campaign):
    return any(k in campaign.lower() for k in GOOGLE_RETENTION_KEYWORDS)


def is_retention_meta(campaign):
    return any(k in campaign.lower() for k in META_RETENTION_KEYWORDS)


# ── Process Google ─────────────────────────────────────────────────────────────

def process_google(gc, mtd_start, mtd_end):
    print("[Google] Reading Google_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Google_Ads")

    brands = defaultdict(lambda: {
        "spend": 0., "impressions": 0,
        "campaigns": set(), "launch_date": None, "last_spend_date": None
    })

    for row in rows:
        campaign = row.get("Campaign", "").strip()
        adgroup  = row.get("Ad_Group", "").strip()
        date_val = row.get("Date", "")

        if not is_retention_google(campaign):
            continue

        brand = extract_google_brand(adgroup)
        if not brand:
            continue

        # Parse date
        d = None
        if isinstance(date_val, str):
            for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
                try:
                    d = datetime.strptime(date_val[:10], fmt).date()
                    break
                except ValueError:
                    continue
        elif hasattr(date_val, 'date'):
            d = date_val.date()
        else:
            d = date_val

        if d is None:
            continue

        try:
            spend = float(row.get("Spend_INR", 0) or 0)
            impr  = int(float(row.get("Impressions", 0) or 0))
        except ValueError:
            continue

        if spend <= 0 and impr <= 0:
            continue

        # Track launch date and last spend date
        if spend > 0:
            if brands[brand]["launch_date"] is None or d < brands[brand]["launch_date"]:
                brands[brand]["launch_date"] = d
            if brands[brand]["last_spend_date"] is None or d > brands[brand]["last_spend_date"]:
                brands[brand]["last_spend_date"] = d

        if not in_range(d, mtd_start, mtd_end):
            continue

        brands[brand]["spend"]       += spend * GOOGLE_META_GST
        brands[brand]["impressions"] += impr
        brands[brand]["campaigns"].add(campaign)

    for b in brands:
        brands[b]["spend"] = round(brands[b]["spend"], 2)

    print(f"  {len(brands)} brands found in Google")
    return brands


def process_google_reach(gc, active_campaigns):
    """Get unique users per brand from ad group level Google Unique Users sheet."""
    print("[Google] Reading Ad Group Unique Users...")
    adgroup_tab = os.getenv("GOOGLE_ADGROUP_UNIQUE_USERS_TAB", "adgroup_unique_users_4")
    rows = read_sheet(gc, GOOGLE_UNIQUE_USERS_SID, adgroup_tab)

    brand_reach = {}
    for row in rows:
        adgroup  = row.get("Ad group", "").strip()
        campaign = row.get("Campaign", "").strip()
        uu       = row.get("Unique users", "").strip()

        # Only retention campaigns
        if not is_retention_google(campaign):
            continue
        if not uu or uu == "--":
            continue

        brand = extract_google_brand(adgroup)
        if not brand:
            continue

        try:
            reach = int(str(uu).replace(",", ""))
        except ValueError:
            continue

        # Take max if brand appears multiple times
        if brand not in brand_reach or reach > brand_reach[brand]:
            brand_reach[brand] = reach

    print(f"  Brand reach from adgroup sheet: {brand_reach}")
    return brand_reach


# ── Process Meta ───────────────────────────────────────────────────────────────

def process_meta(gc, mtd_start, mtd_end):
    print("[Meta] Reading Meta_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Meta_Ads")

    brands = defaultdict(lambda: {
        "spend": 0., "impressions": 0,
        "reach": 0, "campaigns": set(), "launch_date": None, "last_spend_date": None
    })

    for row in rows:
        campaign = row.get("Campaign", "").strip()
        ad       = row.get("Ad", "").strip()
        date_val = row.get("Date", "")

        if not is_retention_meta(campaign):
            continue

        brand = extract_meta_brand(ad)
        if not brand:
            continue

        # Parse date
        d = None
        if isinstance(date_val, str):
            for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
                try:
                    d = datetime.strptime(date_val[:10], fmt).date()
                    break
                except ValueError:
                    continue
        elif hasattr(date_val, 'date'):
            d = date_val.date()
        else:
            d = date_val

        if d is None:
            continue

        try:
            spend = float(row.get("Spend_INR", 0) or 0)
            impr  = int(float(row.get("Impressions", 0) or 0))
            reach = int(float(row.get("Reach", 0) or 0))
        except ValueError:
            continue

        if spend <= 0 and impr <= 0:
            continue

        # Track launch date and last spend date
        if spend > 0:
            if brands[brand]["launch_date"] is None or d < brands[brand]["launch_date"]:
                brands[brand]["launch_date"] = d
            if brands[brand]["last_spend_date"] is None or d > brands[brand]["last_spend_date"]:
                brands[brand]["last_spend_date"] = d

        if not in_range(d, mtd_start, mtd_end):
            continue

        brands[brand]["spend"]       += spend * GOOGLE_META_GST
        brands[brand]["impressions"] += impr
        brands[brand]["reach"]       += reach
        brands[brand]["campaigns"].add(campaign)

    for b in brands:
        brands[b]["spend"] = round(brands[b]["spend"], 2)

    print(f"  {len(brands)} brands found in Meta")
    return brands


def fetch_meta_reach(mtd_start, mtd_end, active_campaigns):
    """Fetch MTD unduplicated reach from Meta API."""
    access_token = os.getenv("META_ACCESS_TOKEN")
    ad_account   = os.getenv("META_AD_ACCOUNT_ID")
    if not access_token or not ad_account:
        return 0
    url    = f"https://graph.facebook.com/v18.0/{ad_account}/insights"
    params = {
        "access_token": access_token,
        "level":        "campaign",
        "fields":       "campaign_name,reach",
        "time_range":   f'{{"since":"{mtd_start}","until":"{mtd_end}"}}',
        "limit":        500,
    }
    total = 0
    try:
        r    = req.get(url, params=params, timeout=30)
        data = r.json().get("data", [])
        for item in data:
            if item.get("campaign_name", "") in active_campaigns:
                total += int(item.get("reach", 0) or 0)
    except Exception as e:
        print(f"[Meta Reach] Error: {e}")
    print(f"  Meta reach: {total}")
    return total


# ── Process Apptrove ───────────────────────────────────────────────────────────

def process_apptrove(gc, mtd_start, mtd_end, google_brands, meta_brands):
    """
    Match Apptrove rows to brands using ad_group (Google) or ad (Meta).
    Sum app_opened per brand.
    """
    print("[Apptrove] Reading Apptrove_MMP...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Apptrove_MMP")

    brand_app = defaultdict(int)

    for row in rows:
        partner  = row.get("partner", "").strip()
        channel  = row.get("channel", "").strip()
        campaign = row.get("campaign", "").strip()
        adgroup  = row.get("ad_group", "").strip()
        ad       = row.get("ad", "").strip()
        date_val = row.get("Date", "")

        if not in_range(date_val, mtd_start, mtd_end):
            continue

        try:
            app_opened = int(row.get("app_opened", 0) or 0)
        except ValueError:
            continue

        if partner == "Google Ads (Adwords)":
            if not channel or channel.strip() == "-":
                continue
            if not is_retention_google(campaign):
                continue
            brand = extract_google_brand(adgroup)
            if brand and brand in google_brands:
                brand_app[brand] += app_opened

        elif partner == "Facebook":
            if not is_retention_meta(campaign):
                continue
            brand = extract_meta_brand(ad)
            if brand and brand in meta_brands:
                brand_app[brand] += app_opened

    print(f"  App traffic computed for {len(brand_app)} brands")
    return brand_app


# ── Merge brands ───────────────────────────────────────────────────────────────

def merge_brands(google_brands, meta_brands):
    merged = {}

    for brand, data in google_brands.items():
        merged[brand] = {
            "spend":           data["spend"],
            "impressions":     data["impressions"],
            "google_reach":    0,
            "meta_reach":      0,
            "campaigns":       data["campaigns"],
            "launch_date":     data["launch_date"],
            "last_spend_date": data["last_spend_date"],
            "source":          "Google",
        }

    for brand, data in meta_brands.items():
        match = None
        for existing in merged:
            if existing.lower() == brand.lower():
                match = existing
                break

        if match:
            merged[match]["spend"]       += data["spend"]
            merged[match]["spend"]        = round(merged[match]["spend"], 2)
            merged[match]["impressions"] += data["impressions"]
            merged[match]["meta_reach"]   = data["reach"]
            merged[match]["campaigns"].update(data["campaigns"])
            merged[match]["source"]       = "Google+Meta"
            if data["launch_date"] and (
                merged[match]["launch_date"] is None or
                data["launch_date"] < merged[match]["launch_date"]
            ):
                merged[match]["launch_date"] = data["launch_date"]
            # Take latest last_spend_date
            if data["last_spend_date"] and (
                merged[match]["last_spend_date"] is None or
                data["last_spend_date"] > merged[match]["last_spend_date"]
            ):
                merged[match]["last_spend_date"] = data["last_spend_date"]
        else:
            merged[brand] = {
                "spend":           data["spend"],
                "impressions":     data["impressions"],
                "google_reach":    0,
                "meta_reach":      data["reach"],
                "campaigns":       data["campaigns"],
                "launch_date":     data["launch_date"],
                "last_spend_date": data["last_spend_date"],
                "source":          "Meta",
            }

    return merged


# ── Build report rows ──────────────────────────────────────────────────────────

def build_rows(merged_brands, brand_app, brand_reach):
    """Build one row per brand sorted by spend desc."""
    now_ist   = datetime.now(IST)
    yesterday = (now_ist.date() - timedelta(days=1))
    rows = []

    for brand, data in sorted(merged_brands.items(), key=lambda x: -x[1]["spend"]):
        impr      = data["impressions"]
        spend     = data["spend"]
        app_open  = brand_app.get(brand, 0)
        ctr       = round(app_open / impr * 100, 2) if impr else "N/A"

        g_reach = brand_reach.get(brand, 0)
        m_reach = data["meta_reach"]
        reach   = max(g_reach, m_reach)

        launch = data["launch_date"].strftime("%d-%m-%Y") if data["launch_date"] else ""

        # Paused date: day after last spend date, only if last spend < yesterday
        last_sd = data["last_spend_date"]
        if last_sd and last_sd < yesterday:
            paused = (last_sd + timedelta(days=1)).strftime("%d-%m-%Y")
        else:
            paused = ""

        rows.append([
            brand,
            launch,
            paused,
            reach,
            impr,
            ctr,
            spend,
            app_open,
            "",   # Avg PPV Before — Mixpanel pending
            "",   # Avg PPV After  — Mixpanel pending
            "",   # Difference     — Mixpanel pending
        ])

    return rows


# ── Write Google Sheet ─────────────────────────────────────────────────────────

def write_report(gc, data_rows, now_ist_str, mtd_start, mtd_end):
    sh = gc.open_by_key(REPORT_SHEET_ID)
    try:
        ws = sh.worksheet("KAM_Brand_Report")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title="KAM_Brand_Report", rows=200, cols=15)

    all_rows = [
        [f"Last Updated: {now_ist_str}"],
        [f"Period: MTD ({mtd_start} → {mtd_end})"],
        ["Note: Reach = Google Unique Users + Meta MTD Reach (campaign level, not brand level)"],
        ["Note: PPV columns blank — Mixpanel integration pending"],
        [],
        HEADERS,
    ]
    all_rows.extend(data_rows)

    ws.clear()
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print(f"[Sheets] Written KAM_Brand_Report — {len(data_rows)} brands")


# ── Write OneDrive Excel ───────────────────────────────────────────────────────

def write_excel(data_rows, now_ist_str, mtd_start, mtd_end):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os as _os

    if not _os.path.exists(ONEDRIVE_PATH):
        print(f"[Excel] File not found: {ONEDRIVE_PATH}")
        return

    wb = load_workbook(ONEDRIVE_PATH)
    if "KAM_Brand_Report" in wb.sheetnames:
        del wb["KAM_Brand_Report"]
    ws = wb.create_sheet("KAM_Brand_Report")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center")

    cr = 1
    ws.cell(row=cr, column=1, value=f"Last Updated: {now_ist_str}").font = Font(italic=True); cr += 1
    ws.cell(row=cr, column=1, value=f"MTD: {mtd_start} → {mtd_end}").font = Font(italic=True); cr += 2

    # Headers
    for ci, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=cr, column=ci, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center
    cr += 1

    # Data
    for row in data_rows:
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=cr, column=ci)
            try:
                cell.value = float(val) if val not in ("", None, "N/A") else val
            except (ValueError, TypeError):
                cell.value = val
        cr += 1

    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = ml + 4

    wb.save(ONEDRIVE_PATH)
    print(f"[Excel] Written KAM_Brand_Report to OneDrive")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("pull_kam_brand_report.py — KAM Brand Report")
    print("=" * 55)

    mtd_start, mtd_end = get_mtd()
    print(f"[Period] MTD: {mtd_start} → {mtd_end}")

    gc          = get_gc()
    now_ist_str = datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S")

    # Fetch data
    google_brands = process_google(gc, mtd_start, mtd_end)
    meta_brands   = process_meta(gc, mtd_start, mtd_end)

    # Reach — adgroup level from Google, daily from Meta
    camp_reach = process_google_reach(gc, set())

    # Merge brands
    merged = merge_brands(google_brands, meta_brands)

    # Apptrove
    brand_app = process_apptrove(gc, mtd_start, mtd_end, google_brands, meta_brands)

    # Build and write
    data_rows = build_rows(merged, brand_app, camp_reach)
    write_report(gc, data_rows, now_ist_str, mtd_start, mtd_end)
    write_excel(data_rows, now_ist_str, mtd_start, mtd_end)

    print(f"\n✅ pull_kam_brand_report.py complete — {len(data_rows)} brands")

    # Print brand mapping for review
    print("\n[Brand Mapping Review]")
    print(f"  {'Ad Name':<45} {'Brand Extracted':<30} {'Source'}")
    print("  " + "-" * 85)
    for brand, data in sorted(merged.items()):
        print(f"  {brand:<45} {brand:<30} {data['source']}")


if __name__ == "__main__":
    main()