"""
pull_team_cost_report.py
Reads raw data from source sheets, computes team-wise cost breakdowns.

Periods:
  Before 12 PM IST: MTD = May 1 → yesterday,  MTD-1 = May 1 → day before yesterday
  After  12 PM IST: MTD = May 1 → today,       MTD-1 = May 1 → yesterday
  Change = MTD - MTD-1 (absolute)

Reads from:
  - GOOGLE_SHEET_ID:    Google_Ads, Meta_Ads, Wati
  - WEBENGAGE_SHEET_ID: Whatsapp_Campaign

Writes to:
  - REPORT_SHEET_ID: Google_Campaigns, Meta_Campaigns, Wati_Cost, WebEngage_Cost, Team_Wise_Cost
  - OneDrive Excel:  Team Wise Cost.xlsx
"""

import os
from datetime import datetime, timedelta
from collections import defaultdict
import pytz
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

IST = pytz.timezone("Asia/Kolkata")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

GOOGLE_SHEET_ID     = os.getenv("GOOGLE_SHEET_ID")
WEBENGAGE_SHEET_ID  = os.getenv("WEBENGAGE_SHEET_ID")
REPORT_SHEET_ID     = os.getenv("REPORT_SHEET_ID")
ONEDRIVE_EXCEL_PATH = os.getenv("ONEDRIVE_EXCEL_PATH", r"C:\Users\anitb\OneDrive - Agrim Wholesale Private Limited\Marketing_Reports\Team Wise Cost.xlsx")

TEAMS = ["Customer Support", "Superstar", "Marketing Onboarding", "Marketing Retention", "Voice AI", "Supply"]

WATI_COST_RATE      = 0.96
WEBENGAGE_BASE_RATE = 0.87
WEBENGAGE_GST       = 1.18
GOOGLE_META_GST     = 1.18


# ── Date periods ───────────────────────────────────────────────────────────────

def get_periods():
    now_ist   = datetime.now(IST)
    today     = now_ist.date()
    yesterday = today - timedelta(days=1)
    mtd_start = today.replace(day=1)

    if now_ist.hour < 12:
        mtd_end  = yesterday
        mtd1_end = yesterday - timedelta(days=1)
    else:
        mtd_end  = today
        mtd1_end = yesterday

    return mtd_start, mtd_end, mtd1_end


def in_period(date_val, start, end):
    if isinstance(date_val, datetime):
        d = date_val.date()
    elif isinstance(date_val, str):
        for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
            try:
                d = datetime.strptime(date_val[:10], fmt).date()
                break
            except ValueError:
                continue
        else:
            return False
    else:
        d = date_val
    return start <= d <= end


# ── Auth ───────────────────────────────────────────────────────────────────────

def get_gc():
    creds_path = (
        os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH")
        or os.getenv("GOOGLE_CREDENTIALS_PATH")
        or "service_account.json"
    )
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


# ── Team mapping ───────────────────────────────────────────────────────────────

def map_team_wati(campaign_name):
    n = campaign_name.lower().strip()
    if n == "voice_ai":                                          return "Voice AI"
    if n.startswith("ss_"):                                      return "Superstar"
    if "customer_support" in n or n.startswith("bot_") or n == "test": return "Customer Support"
    if "marketing_onboarding" in n:                              return "Marketing Onboarding"
    if "marketing_retention" in n:                               return "Marketing Retention"
    if "marketing" in n:                                         return "Marketing Retention"
    return "Marketing Retention"


def map_team_meta(campaign_name):
    n = campaign_name.lower().strip()
    if "superstar" in n or n.startswith("ss_"):                  return "Superstar"
    if "seller" in n:                                            return "Supply"
    if n.startswith("ar_") and "install" in n:                   return "Marketing Onboarding"
    if any(k in n for k in ["onboarding", "aci", "otp_entered", "kyced_but_not_transacted"]): return "Marketing Onboarding"
    if any(k in n for k in ["retention", "ace"]):                return "Marketing Retention"
    if "marketing" in n:                                         return "Marketing Retention"
    return "Marketing Retention"


def map_team_google(campaign_name):
    n = campaign_name.lower().strip()
    if any(k in n for k in ["aci", "otp_entered", "kyced_but_not_transacted", "install", "onboarding"]): return "Marketing Onboarding"
    if any(k in n for k in ["retention", "ace"]):                return "Marketing Retention"
    if "marketing" in n:                                         return "Marketing Retention"
    return "Marketing Retention"


def map_team_webengage(campaign_type, campaign_name):
    t = str(campaign_type).strip().lower()
    n = campaign_name.lower().strip()
    if t == "relay":       return "Marketing Retention"
    if t == "journey":     return "Marketing Retention" if "add to cart but not purchased" in n else "Marketing Onboarding"
    if t == "one time":    return "Marketing Onboarding" if "onboarding" in n else "Marketing Retention"
    return "Marketing Retention"


# ── Read sheet ─────────────────────────────────────────────────────────────────

def read_sheet(gc, sheet_id, tab_name):
    sh   = gc.open_by_key(sheet_id)
    ws   = sh.worksheet(tab_name)
    rows = ws.get_all_values()
    if not rows or len(rows) < 2:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


# ── Process Google Ads ─────────────────────────────────────────────────────────

def process_google(gc, mtd_start, mtd_end, mtd1_end):
    print("[Google] Reading Google_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Google_Ads")

    data = defaultdict(lambda: {"mtd": 0, "mtd1": 0, "team": ""})

    for row in rows:
        campaign = row.get("Campaign", "").strip()
        if not campaign:
            continue
        try:
            spend = float(row.get("Spend_INR", 0) or 0) * GOOGLE_META_GST
        except ValueError:
            continue
        date_val = row.get("Date", "")
        team     = map_team_google(campaign)
        data[campaign]["team"] = team

        if in_period(date_val, mtd_start, mtd_end):
            data[campaign]["mtd"] += spend
        if in_period(date_val, mtd_start, mtd1_end):
            data[campaign]["mtd1"] += spend

    output = []
    for campaign, vals in sorted(data.items(), key=lambda x: -x[1]["mtd"]):
        mtd  = round(vals["mtd"], 2)
        mtd1 = round(vals["mtd1"], 2)
        output.append([campaign, vals["team"], mtd, mtd1, round(mtd - mtd1, 2)])
    print(f"[Google] {len(output)} campaigns processed")
    return output


# ── Process Meta Ads ───────────────────────────────────────────────────────────

def process_meta(gc, mtd_start, mtd_end, mtd1_end):
    print("[Meta] Reading Meta_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Meta_Ads")

    data = defaultdict(lambda: {"mtd": 0, "mtd1": 0, "team": ""})

    for row in rows:
        campaign = row.get("Campaign", "").strip()
        if not campaign:
            continue
        try:
            spend = float(row.get("Spend_INR", 0) or 0) * GOOGLE_META_GST
        except ValueError:
            continue
        date_val = row.get("Date", "")
        team     = map_team_meta(campaign)
        data[campaign]["team"] = team

        if in_period(date_val, mtd_start, mtd_end):
            data[campaign]["mtd"] += spend
        if in_period(date_val, mtd_start, mtd1_end):
            data[campaign]["mtd1"] += spend

    output = []
    for campaign, vals in sorted(data.items(), key=lambda x: -x[1]["mtd"]):
        mtd  = round(vals["mtd"], 2)
        mtd1 = round(vals["mtd1"], 2)
        output.append([campaign, vals["team"], mtd, mtd1, round(mtd - mtd1, 2)])
    print(f"[Meta] {len(output)} campaigns processed")
    return output


# ── Process Wati ───────────────────────────────────────────────────────────────

def process_wati(gc, mtd_start, mtd_end, mtd1_end):
    print("[Wati] Reading Wati...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Wati")

    data = defaultdict(lambda: {"mtd": 0, "mtd1": 0})

    for row in rows:
        campaign = row.get("Campaign_Name", "").strip()
        if not campaign:
            continue
        try:
            sent   = float(row.get("Sent", 0) or 0)
            failed = float(row.get("Failed", 0) or 0)
        except ValueError:
            continue
        cost     = (sent - failed) * WATI_COST_RATE
        date_val = row.get("Date", "")
        team     = map_team_wati(campaign)

        if in_period(date_val, mtd_start, mtd_end):
            data[team]["mtd"] += cost
        if in_period(date_val, mtd_start, mtd1_end):
            data[team]["mtd1"] += cost

    wati_teams = ["Customer Support", "Superstar", "Marketing Onboarding", "Marketing Retention", "Voice AI"]
    output = []
    for team in wati_teams:
        mtd  = round(data[team]["mtd"], 2)
        mtd1 = round(data[team]["mtd1"], 2)
        output.append([team, mtd, mtd1, round(mtd - mtd1, 2)])
    print(f"[Wati] {len(output)} teams processed")
    return output


# ── Process WebEngage ──────────────────────────────────────────────────────────

def process_webengage(gc):
    print("[WebEngage] Reading Whatsapp_Campaign...")
    rows = read_sheet(gc, WEBENGAGE_SHEET_ID, "Whatsapp_Campaign")

    now_ist       = datetime.now(IST)
    current_month = now_ist.month
    current_year  = now_ist.year
    data          = defaultdict(float)

    for row in rows:
        campaign_name = row.get("Campaign Name", "").strip()
        campaign_type = row.get("Type of Campaign", "").strip()
        if not campaign_name:
            continue
        start_str  = row.get("Reporting Period Start Date", "").strip()
        start_date = None
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                start_date = datetime.strptime(start_str[:10], fmt)
                break
            except ValueError:
                continue
        if start_date is None:
            continue
        if start_date.month != current_month or start_date.year != current_year:
            continue
        try:
            delivered = float(row.get("Delivered", 0) or 0)
        except ValueError:
            continue
        cost = delivered * WEBENGAGE_BASE_RATE * WEBENGAGE_GST
        team = map_team_webengage(campaign_type, campaign_name)
        data[team] += cost

    we_teams = ["Marketing Onboarding", "Marketing Retention"]
    output = []
    for team in we_teams:
        output.append([team, round(data[team], 2)])
    print(f"[WebEngage] {len(output)} teams processed")
    return output


# ── Build Team Wise Cost tables ────────────────────────────────────────────────

def build_team_wise_cost(google_rows, meta_rows, wati_rows, webengage_rows,
                          mtd_start, mtd_end, mtd1_end):
    col_teams = ["Superstar", "Marketing Onboarding", "Marketing Retention", "Voice AI", "Supply"]

    # Build spend indexes: period → team → spend
    google_mtd  = defaultdict(float)
    google_mtd1 = defaultdict(float)
    for row in google_rows:
        # [campaign, team, mtd, mtd1, change]
        team = row[1]
        google_mtd[team]  += row[2]
        google_mtd1[team] += row[3]

    meta_mtd  = defaultdict(float)
    meta_mtd1 = defaultdict(float)
    for row in meta_rows:
        team = row[1]
        meta_mtd[team]  += row[2]
        meta_mtd1[team] += row[3]

    wati_mtd  = defaultdict(float)
    wati_mtd1 = defaultdict(float)
    for row in wati_rows:
        # [team, mtd, mtd1, change]
        team = row[0]
        wati_mtd[team]  += row[1]
        wati_mtd1[team] += row[2]

    we_idx = defaultdict(float)
    for row in webengage_rows:
        we_idx[row[0]] += row[1]

    def build_table(label, g_idx, m_idx, w_idx, include_webengage=False):
        rows_out = []
        rows_out.append([label, "Superstar", "Marketing Onboarding", "Marketing Retention", "Voice AI", "Supply"])
        for channel, idx in [("Google", g_idx), ("Meta", m_idx), ("WhatsApp", None)]:
            if channel == "WhatsApp":
                row = ["WhatsApp"]
                for team in col_teams:
                    wati_cost = w_idx.get(team, 0)
                    we_cost   = we_idx.get(team, 0) if include_webengage else 0
                    row.append(round(wati_cost + we_cost, 2))
            else:
                row = [channel]
                for team in col_teams:
                    row.append(round(idx.get(team, 0), 2))
            rows_out.append(row)

        total_row = ["Total"]
        for i in range(len(col_teams)):
            col_idx = i + 1
            total_row.append(round(sum(r[col_idx] for r in rows_out[1:]), 2))
        rows_out.append(total_row)
        return rows_out

    def build_change_table(mtd_table, mtd1_table):
        rows_out = []
        rows_out.append(["Change (MTD - MTD-1)"] + [""] * 5)
        for r_mtd, r_mtd1 in zip(mtd_table[1:], mtd1_table[1:]):
            change_row = [r_mtd[0]]
            for i in range(1, len(r_mtd)):
                try:
                    change_row.append(round(float(r_mtd[i]) - float(r_mtd1[i]), 2))
                except (ValueError, TypeError):
                    change_row.append("")
            rows_out.append(change_row)
        return rows_out

    mtd_label  = f"MTD ({mtd_start.strftime('%d-%m-%Y')} → {mtd_end.strftime('%d-%m-%Y')})"
    mtd1_label = f"MTD-1 ({mtd_start.strftime('%d-%m-%Y')} → {mtd1_end.strftime('%d-%m-%Y')})"

    mtd_table  = build_table(mtd_label,  google_mtd,  meta_mtd,  wati_mtd,  include_webengage=True)
    mtd1_table = build_table(mtd1_label, google_mtd1, meta_mtd1, wati_mtd1, include_webengage=False)
    chg_table  = build_change_table(mtd_table, mtd1_table)

    all_rows = []
    all_rows += mtd_table
    all_rows.append([""] * 6)
    all_rows += mtd1_table
    all_rows.append([""] * 6)
    all_rows += chg_table

    return all_rows


# ── Write to report sheet ──────────────────────────────────────────────────────

def write_tab(sh, tab_name, headers, data_rows):
    try:
        ws = sh.worksheet(tab_name)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=tab_name, rows=1000, cols=20)

    now_ist  = datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S")
    all_rows = [[f"Last Updated: {now_ist}"] + [""] * (len(headers) - 1)]
    all_rows.append(headers)
    all_rows.extend(data_rows)

    ws.clear()
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print(f"[Sheets] Written {len(data_rows)} rows to '{tab_name}'")


def write_team_wise_cost(sh, table_rows):
    tab_name = "Team_Wise_Cost"
    try:
        ws = sh.worksheet(tab_name)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=tab_name, rows=100, cols=10)

    now_ist  = datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S")
    all_rows = [[f"Last Updated: {now_ist}"], []]
    all_rows.extend(table_rows)

    ws.clear()
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print(f"[Sheets] Written Team_Wise_Cost tab")


# ── Write to OneDrive Excel ────────────────────────────────────────────────────

def write_excel(gc, now_ist_str):
    print("[Excel] Reading Team_Wise_Cost from Google Sheet...")
    ws       = gc.open_by_key(REPORT_SHEET_ID).worksheet("Team_Wise_Cost")
    all_rows = ws.get_all_values()

    # Extract all 3 tables: MTD, MTD-1, Change
    table_labels   = []
    extracted      = {}
    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        if row and row[0] and any(k in row[0] for k in ["MTD", "Change"]):
            label      = row[0]
            table_rows = [row]
            i += 1
            while i < len(all_rows):
                r = all_rows[i]
                if not any(r):
                    break
                table_rows.append(r)
                i += 1
            extracted[label] = table_rows
            table_labels.append(label)
        else:
            i += 1

    from openpyxl import load_workbook
    import os as _os
    if _os.path.exists(ONEDRIVE_EXCEL_PATH):
        wb = load_workbook(ONEDRIVE_EXCEL_PATH)
        if "Team Wise Cost" in wb.sheetnames:
            del wb["Team Wise Cost"]
        ws_xl = wb.create_sheet("Team Wise Cost", 0)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb    = Workbook()
        ws_xl = wb.active
        ws_xl.title = "Team Wise Cost"

    label_fill  = PatternFill("solid", fgColor="2E75B6")
    label_font  = Font(bold=True, color="FFFFFF")
    total_font  = Font(bold=True)
    center      = Alignment(horizontal="center")

    current_row = 1
    ws_xl.cell(row=current_row, column=1, value=f"Last Updated: {now_ist_str}").font = Font(italic=True)
    current_row += 2

    for label in table_labels:
        if label not in extracted:
            continue
        for row_idx, row in enumerate(extracted[label]):
            for col_idx, val in enumerate(row, start=1):
                cell = ws_xl.cell(row=current_row, column=col_idx)
                try:
                    cell.value = float(val) if val not in ("", None) else ""
                except (ValueError, TypeError):
                    cell.value = val
                if row_idx == 0:
                    cell.font  = label_font
                    cell.fill  = label_fill
                    cell.alignment = center
                elif row and row[0] == "Total":
                    cell.font = total_font
            current_row += 1
        current_row += 1

    for col in ws_xl.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws_xl.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(ONEDRIVE_EXCEL_PATH)
    print(f"[Excel] Written to {ONEDRIVE_EXCEL_PATH}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("pull_team_cost_report.py — Daily Cost Report")
    print("=" * 55)

    mtd_start, mtd_end, mtd1_end = get_periods()
    print(f"[Periods] MTD: {mtd_start} → {mtd_end} | MTD-1: {mtd_start} → {mtd1_end}")

    gc        = get_gc()
    report_sh = gc.open_by_key(REPORT_SHEET_ID)

    google_rows    = process_google(gc, mtd_start, mtd_end, mtd1_end)
    meta_rows      = process_meta(gc, mtd_start, mtd_end, mtd1_end)
    wati_rows      = process_wati(gc, mtd_start, mtd_end, mtd1_end)
    webengage_rows = process_webengage(gc)

    write_tab(report_sh, "Google_Campaigns",
              ["Campaign", "Team", "MTD_Spend", "MTD1_Spend", "Change_Spend"],
              google_rows)
    write_tab(report_sh, "Meta_Campaigns",
              ["Campaign", "Team", "MTD_Spend", "MTD1_Spend", "Change_Spend"],
              meta_rows)
    write_tab(report_sh, "Wati_Cost",
              ["Team", "MTD_Cost", "MTD1_Cost", "Change_Cost"],
              wati_rows)
    write_tab(report_sh, "WebEngage_Cost",
              ["Team", "MTD_Cost"],
              webengage_rows)

    team_wise_rows = build_team_wise_cost(
        google_rows, meta_rows, wati_rows, webengage_rows,
        mtd_start, mtd_end, mtd1_end
    )
    write_team_wise_cost(report_sh, team_wise_rows)

    now_ist_str = datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S")
    write_excel(gc, now_ist_str)

    print("\n✅ pull_team_cost_report.py complete")


if __name__ == "__main__":
    main()