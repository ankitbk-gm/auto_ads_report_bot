"""
pull_prev_month_report.py
Consolidated Previous Month Report — 3 sections in one tab.

Data sources (all available for previous month):
  Google_Ads, Meta_Ads, Apptrove_MMP, Wati  — 90-day rolling
  Whatsapp_Campaign (WebEngage)              — full month rows available

Sections:
  1. Team Wise Cost   — Google / Meta / Wati / WebEngage by team
  2. Onboarding       — KYC + First TXN (Wati + WebEngage combined for WhatsApp)
  3. Retention        — Dynamic segments (Brand, Subcat + any others)

Writes to:
  REPORT_SHEET_ID → Prev_Month_Report tab
  OneDrive Excel  → Team Wise Cost.xlsx → Prev_Month_Report sheet
"""

import os
import time
from datetime import datetime, timedelta
from collections import defaultdict
import pytz
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials
import requests as req

load_dotenv()

IST    = pytz.timezone("Asia/Kolkata")
SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive"]

GOOGLE_SHEET_ID         = os.getenv("GOOGLE_SHEET_ID")
WEBENGAGE_SHEET_ID      = os.getenv("WEBENGAGE_SHEET_ID")
REPORT_SHEET_ID         = os.getenv("REPORT_SHEET_ID")
GOOGLE_UNIQUE_USERS_SID = os.getenv("GOOGLE_UNIQUE_USERS_SHEET_ID")
GOOGLE_UNIQUE_USERS_TAB = os.getenv("GOOGLE_UNIQUE_USERS_TAB", "campaign_unique")
ONEDRIVE_PATH           = os.getenv("ONEDRIVE_EXCEL_PATH",
    r"C:\Users\anitb\OneDrive - Agrim Wholesale Private Limited\Marketing_Reports\Team Wise Cost.xlsx")

REPORT_TAB          = "Prev_Month_Report"
GOOGLE_META_GST     = 1.18
WATI_COST_RATE      = 0.96
WEBENGAGE_BASE_RATE = 0.87
WEBENGAGE_GST       = 1.18
COST_TEAMS = ["Superstar", "Marketing Onboarding", "Marketing Retention", "Voice AI", "Supply"]

# ── Classifiers ─────────────────────────────────────────────────────────────────
WATI_KYC_KEYWORDS   = ["otp_entered_not_kyc", "marketing_install"]
GOOGLE_RETENTION_KW = ["ace", "ar_purchasers"]
META_RETENTION_KW   = ["retention"]

def is_google_kyc(c): return any(k in c.lower() for k in ["otp_entered", "aci"])
def is_google_txn(c): return "kyced_but_not_transacted" in c.lower()
def is_meta_kyc(c):
    n = c.lower(); return any(k in n for k in ["install","otp_entered"]) and n.startswith("ar_onboarding")
def is_meta_txn(c):
    n = c.lower(); return "kyc" in n and n.startswith("ar_onboarding")
def is_wati_kyc(c): return any(k in c.lower() for k in WATI_KYC_KEYWORDS)
def is_wati_ob(c):  return "marketing_onboarding" in c.lower()
def is_we_kyc(j):   return "kyc_journey" in j.lower()
def is_we_ob(t, n):
    t = t.lower().strip(); n = n.lower().strip()
    if t == "relay":    return False
    if t == "journey":  return "add to cart but not purchased" not in n
    if t == "one time": return "onboarding" in n
    return False

def is_retention_google(c): return any(k in c.lower() for k in GOOGLE_RETENTION_KW)
def is_retention_meta(c):   return any(k in c.lower() for k in META_RETENTION_KW)

def classify_google_adgroup(ag):
    n = ag.lower()
    if "brand" in n:   return "Brand"
    if "categor" in n: return "Subcat"
    return ag.strip() if ag.strip() else "Other"

def classify_meta_ad(ad):
    n = ad.lower()
    if "brand" in n:                           return "Brand"
    if "category" in n or "categor" in n:      return "Subcat"
    if "catalogue" in n or "gibberellic" in n: return "Subcat"
    return ad.strip() if ad.strip() else "Other"

def get_segments(d):
    def o(k): return (0,k) if k=="Brand" else (1,k) if k=="Subcat" else (2,k)
    return sorted(d.keys(), key=o)

def map_team_google(c):
    n = c.lower()
    if any(k in n for k in ["aci","otp_entered","kyced_but_not_transacted","install","onboarding"]): return "Marketing Onboarding"
    if any(k in n for k in ["retention","ace"]): return "Marketing Retention"
    return "Marketing Retention"

def map_team_meta(c):
    n = c.lower()
    if "superstar" in n or n.startswith("ss_"):                                               return "Superstar"
    if "seller" in n:                                                                          return "Supply"
    if n.startswith("ar_") and "install" in n:                                                return "Marketing Onboarding"
    if any(k in n for k in ["onboarding","aci","otp_entered","kyced_but_not_transacted"]):    return "Marketing Onboarding"
    if any(k in n for k in ["retention","ace"]):                                              return "Marketing Retention"
    return "Marketing Retention"

def map_team_wati(c):
    n = c.lower()
    if n == "voice_ai":                                                  return "Voice AI"
    if n.startswith("ss_"):                                              return "Superstar"
    if "customer_support" in n or n.startswith("bot_") or n == "test":  return "Customer Support"
    if "marketing_onboarding" in n:                                      return "Marketing Onboarding"
    return "Marketing Retention"

def map_team_webengage(ct, cn):
    t = ct.lower().strip(); n = cn.lower().strip()
    if t == "relay":    return "Marketing Retention"
    if t == "journey":  return "Marketing Retention" if "add to cart but not purchased" in n else "Marketing Onboarding"
    if t == "one time": return "Marketing Onboarding" if "onboarding" in n else "Marketing Retention"
    return "Marketing Retention"


# ── Auth ────────────────────────────────────────────────────────────────────────
def get_gc():
    p = (os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH")
         or os.getenv("GOOGLE_CREDENTIALS_PATH") or "service_account.json")
    return gspread.authorize(Credentials.from_service_account_file(p, scopes=SCOPES))

def get_prev_month_range():
    today    = datetime.now(IST).date()
    prev_end = today.replace(day=1) - timedelta(days=1)
    prev_st  = prev_end.replace(day=1)
    return prev_st, prev_end, prev_st.strftime("%B %Y")

def read_sheet(gc, sid, tab, retries=4, backoff=10):
    for attempt in range(1, retries+1):
        try:
            rows = gc.open_by_key(sid).worksheet(tab).get_all_values()
            if not rows or len(rows) < 2: return []
            h = rows[0]; return [dict(zip(h, r)) for r in rows[1:]]
        except Exception as e:
            if attempt < retries and any(c in str(e) for c in ["503","500","429","502"]):
                time.sleep(backoff * attempt)
            else: raise

def in_range(dv, s, e):
    if isinstance(dv, str):
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try: d = datetime.strptime(dv[:10], fmt).date(); return s <= d <= e
            except: continue
        return False
    if hasattr(dv, 'date'): return s <= dv.date() <= e
    return s <= dv <= e

def we_in_prev_month(start_str, ps, pe):
    """Check if WebEngage reporting period start date falls in prev month."""
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(start_str[:10], fmt).date()
            return ps <= d <= pe
        except: continue
    return False

def sdiv(a, b): return "N/A" if not b else round(a/b*100, 2)
def sctr(cl, im): return "N/A" if not im else round(cl/im*100, 2)
def sfreq(im, r): return "N/A" if not r else round(im/r, 2)
def cpx(sp, cv): return "N/A" if not cv else round(sp/cv, 2)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — TEAM WISE COST
# ══════════════════════════════════════════════════════════════════════════════
def build_cost_section(google_rows, meta_rows, wati_rows, we_rows, ps, pe, label):
    print("[Cost] Processing...")
    g = defaultdict(float); m = defaultdict(float); w = defaultdict(float); we = defaultdict(float)

    for row in google_rows:
        c = row.get("Campaign","").strip()
        if not c: continue
        try: sp = float(row.get("Spend_INR",0) or 0) * GOOGLE_META_GST
        except: continue
        if sp > 0 and in_range(row.get("Date",""), ps, pe):
            g[map_team_google(c)] += sp

    for row in meta_rows:
        c = row.get("Campaign","").strip()
        if not c: continue
        try: sp = float(row.get("Spend_INR",0) or 0) * GOOGLE_META_GST
        except: continue
        if sp > 0 and in_range(row.get("Date",""), ps, pe):
            m[map_team_meta(c)] += sp

    for row in wati_rows:
        c = row.get("Campaign_Name","").strip()
        if not c: continue
        try: sent = float(row.get("Sent",0) or 0); fail = float(row.get("Failed",0) or 0)
        except: continue
        if in_range(row.get("Date",""), ps, pe):
            w[map_team_wati(c)] += (sent - fail) * WATI_COST_RATE

    for row in we_rows:
        cn = row.get("Campaign Name","").strip()
        ct = row.get("Type of Campaign","").strip()
        if not cn: continue
        if not we_in_prev_month(row.get("Reporting Period Start Date",""), ps, pe): continue
        try: dlvd = float(row.get("Delivered",0) or 0)
        except: continue
        cost = dlvd * WEBENGAGE_BASE_RATE * WEBENGAGE_GST
        we[map_team_webengage(ct, cn)] += cost

    def mrow(ch, idx):
        vals = [round(idx.get(t,0), 2) for t in COST_TEAMS]
        return [ch] + vals + [round(sum(vals), 2)]

    # WhatsApp = Wati + WebEngage combined
    wa = defaultdict(float)
    for team in COST_TEAMS:
        wa[team] = round(w.get(team,0) + we.get(team,0), 2)

    rows = [mrow("Google", g), mrow("Meta", m), mrow("WhatsApp (Wati+WE)", wa)]
    tot  = ["Total"]
    for i in range(len(COST_TEAMS)):
        tot.append(round(sum(r[i+1] for r in rows), 2))
    tot.append(round(sum(tot[1:]), 2))
    rows.append(tot)

    hdr = ["Channel"] + COST_TEAMS + ["Total"]
    print(f"  G:{round(sum(g.values()),2)} M:{round(sum(m.values()),2)} "
          f"Wati:{round(sum(w.values()),2)} WE:{round(sum(we.values()),2)}")
    return [[f"1. TEAM WISE COST — {label}"], hdr] + rows


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — ONBOARDING
# ══════════════════════════════════════════════════════════════════════════════
def build_onboarding_section(google_rows, meta_rows, wati_rows, we_rows, apptrove_rows, unique_rows,
                             gc, ps, pe, label):
    print("[Onboarding] Processing...")
    ps_str = ps.strftime("%Y-%m-%d"); pe_str = pe.strftime("%Y-%m-%d")

    # Google
    gd = {"kyc_sp":0.,"txn_sp":0.,"kyc_im":0,"txn_im":0,"kyc_cl":0,"txn_cl":0,
          "kyc_camps":set(),"txn_camps":set()}
    for row in google_rows:
        c = row.get("Campaign","").strip()
        if not c or not in_range(row.get("Date",""), ps, pe): continue
        try: sp=float(row.get("Spend_INR",0) or 0)*GOOGLE_META_GST; im=int(float(row.get("Impressions",0) or 0)); cl=int(float(row.get("Clicks",0) or 0))
        except: continue
        if is_google_kyc(c): gd["kyc_sp"]+=sp;gd["kyc_im"]+=im;gd["kyc_cl"]+=cl;gd["kyc_camps"].add(c)
        elif is_google_txn(c): gd["txn_sp"]+=sp;gd["txn_im"]+=im;gd["txn_cl"]+=cl;gd["txn_camps"].add(c)
    gd["kyc_sp"]=round(gd["kyc_sp"],2); gd["txn_sp"]=round(gd["txn_sp"],2)

    # Google reach
    g_kyc_r = g_txn_r = 0
    try:
        for row in unique_rows:
            c = row.get("Campaign","").strip(); uu = row.get("Unique users","").strip()
            if not uu or uu == "--": continue
            try: v = int(str(uu).replace(",",""))
            except: continue
            if c in gd["kyc_camps"]: g_kyc_r += v
            elif c in gd["txn_camps"]: g_txn_r += v
    except Exception as e: print(f"  [GReach] {e}")

    # Meta
    md = {"kyc_sp":0.,"txn_sp":0.,"kyc_im":0,"txn_im":0,"kyc_cl":0,"txn_cl":0,
          "kyc_camps":set(),"txn_camps":set()}
    for row in meta_rows:
        c = row.get("Campaign","").strip()
        if not c or not in_range(row.get("Date",""), ps, pe): continue
        try: sp=float(row.get("Spend_INR",0) or 0)*GOOGLE_META_GST; im=int(float(row.get("Impressions",0) or 0)); cl=int(float(row.get("Clicks",0) or 0))
        except: continue
        if is_meta_kyc(c): md["kyc_sp"]+=sp;md["kyc_im"]+=im;md["kyc_cl"]+=cl;md["kyc_camps"].add(c)
        elif is_meta_txn(c): md["txn_sp"]+=sp;md["txn_im"]+=im;md["txn_cl"]+=cl;md["txn_camps"].add(c)
    md["kyc_sp"]=round(md["kyc_sp"],2); md["txn_sp"]=round(md["txn_sp"],2)

    # Meta API
    mk_r=mt_r=mk_im=mt_im=mk_cl=mt_cl=0
    at=os.getenv("META_ACCESS_TOKEN"); aa=os.getenv("META_AD_ACCOUNT_ID")
    if at and aa:
        try:
            url = f"https://graph.facebook.com/v18.0/{aa}/insights"
            params = {"access_token":at,"level":"campaign","fields":"campaign_name,reach,impressions,clicks",
                      "time_range":f'{{"since":"{ps_str}","until":"{pe_str}"}}','limit':500}
            for item in req.get(url, params=params, timeout=30).json().get("data",[]):
                cn=item.get("campaign_name","").strip()
                r=int(item.get("reach",0) or 0); im=int(item.get("impressions",0) or 0); cl=int(item.get("clicks",0) or 0)
                if cn in md["kyc_camps"]: mk_r+=r;mk_im+=im;mk_cl+=cl
                elif cn in md["txn_camps"]: mt_r+=r;mt_im+=im;mt_cl+=cl
        except Exception as e: print(f"  [MetaAPI] {e}")

    # WhatsApp = Wati + WebEngage combined
    wk_sent=wk_cost=wt_sent=wt_cost=0.
    # Wati
    for row in wati_rows:
        c = row.get("Campaign_Name","").strip()
        if not c or not in_range(row.get("Date",""), ps, pe): continue
        try: sent=int(row.get("Sent",0) or 0); dlvd=int(row.get("Delivered",0) or 0)
        except: continue
        cost = dlvd * WATI_COST_RATE
        if is_wati_kyc(c): wk_sent+=sent; wk_cost+=cost
        elif is_wati_ob(c): wt_sent+=sent; wt_cost+=cost
    # WebEngage
    for row in we_rows:
        cn = row.get("Campaign Name","").strip()
        ct = row.get("Type of Campaign","").strip()
        jn = row.get("Journey Name","").strip()
        if not cn or not we_in_prev_month(row.get("Reporting Period Start Date",""), ps, pe): continue
        try: sent=int(row.get("Sent",0) or 0); dlvd=int(row.get("Delivered",0) or 0)
        except: continue
        cost = dlvd * WEBENGAGE_BASE_RATE * WEBENGAGE_GST
        if is_we_kyc(jn): wk_sent+=sent; wk_cost+=cost
        elif is_we_ob(ct, cn) and not is_we_kyc(jn): wt_sent+=sent; wt_cost+=cost
    wk_cost=round(wk_cost,2); wt_cost=round(wt_cost,2)

    # Apptrove
    WA_KW = ["webengage", "whatsapp", "wati", "wa_", "wa bulk"]
    app = {"g_kyc_fhpv":0,"g_txn_fps":0,"m_kyc_fhpv":0,"m_txn_fps":0,"wa_kyc_fhpv":0,"wa_txn_fps":0}
    for row in apptrove_rows:
        partner=row.get("partner","").strip()
        channel=row.get("channel","").strip()
        if not in_range(row.get("Date",""), ps, pe): continue
        try: fhpv=int(row.get("first_homePage_viewed",0) or 0); fps=int(row.get("first_purchase_success",0) or 0)
        except: fhpv=fps=0
        if partner == "Google Ads (Adwords)":
            if not channel or channel.strip() == "-": continue
            app["g_kyc_fhpv"] += fhpv
            if row.get("campaign","").strip() in gd["txn_camps"]:
                app["g_txn_fps"] += fps
        elif partner == "Facebook":
            app["m_kyc_fhpv"] += fhpv
            if row.get("campaign","").strip() in md["txn_camps"]:
                app["m_txn_fps"] += fps
        else:
            pl=partner.lower(); ch=channel.lower()
            if any(k in pl or k in ch for k in WA_KW):
                app["wa_kyc_fhpv"]+=fhpv; app["wa_txn_fps"]+=fps

    print(f"  KYC FHPV G:{app['g_kyc_fhpv']} M:{app['m_kyc_fhpv']} | TXN FPS G:{app['g_txn_fps']} M:{app['m_txn_fps']}")
    print(f"  WA KYC sent:{wk_sent} cost:{wk_cost} | WA TXN sent:{wt_sent} cost:{wt_cost}")

    kh = ["Channel","Reach","Impressions","CTR (%)","Frequency","First Home Page View","Amount Spent","Cost Per KYC"]
    th = ["Channel","Reach","Impressions","CTR (%)","Frequency","First Purchase Success","Amount Spent","Cost Per TXN"]

    kyc = [
        [f"KYC — {label}"], kh,
        ["Google",   g_kyc_r, gd["kyc_im"], sctr(gd["kyc_cl"],gd["kyc_im"]), sfreq(gd["kyc_im"],g_kyc_r), app["g_kyc_fhpv"], gd["kyc_sp"], cpx(gd["kyc_sp"],app["g_kyc_fhpv"])],
        ["Meta",     mk_r,    mk_im,         sctr(mk_cl,mk_im),               sfreq(mk_im,mk_r),            app["m_kyc_fhpv"], md["kyc_sp"], cpx(md["kyc_sp"],app["m_kyc_fhpv"])],
        ["WhatsApp", wk_sent, "N/A",         "N/A",                           "N/A",                        app["wa_kyc_fhpv"],wk_cost,      cpx(wk_cost,app["wa_kyc_fhpv"])],
    ]
    txn = [
        [f"First TXN — {label}"], th,
        ["Google",   g_txn_r, gd["txn_im"], sctr(gd["txn_cl"],gd["txn_im"]), sfreq(gd["txn_im"],g_txn_r), app["g_txn_fps"], gd["txn_sp"], cpx(gd["txn_sp"],app["g_txn_fps"])],
        ["Meta",     mt_r,    mt_im,         sctr(mt_cl,mt_im),               sfreq(mt_im,mt_r),            app["m_txn_fps"], md["txn_sp"], cpx(md["txn_sp"],app["m_txn_fps"])],
        ["WhatsApp", wt_sent, "N/A",         "N/A",                           "N/A",                        app["wa_txn_fps"],wt_cost,      cpx(wt_cost,app["wa_txn_fps"])],
    ]
    return [[f"2. ONBOARDING — {label}"]] + kyc + [[]] + txn


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — RETENTION (DYNAMIC SEGMENTS)
# ══════════════════════════════════════════════════════════════════════════════
def build_retention_section(google_rows, meta_rows, apptrove_rows, unique_rows,
                             gc, ps, pe, label):
    print("[Retention] Processing...")
    ps_str = ps.strftime("%Y-%m-%d"); pe_str = pe.strftime("%Y-%m-%d")

    def esp(): return {"spend":0.,"impressions":0}
    def eap(): return {"app_opened":0,"purchase":0}

    g_spend = defaultdict(esp); g_active = set()
    for row in google_rows:
        c=row.get("Campaign","").strip(); ag=row.get("Ad_Group","").strip()
        if not is_retention_google(c) or not in_range(row.get("Date",""), ps, pe): continue
        cls=classify_google_adgroup(ag)
        try: sp=float(row.get("Spend_INR",0) or 0)*GOOGLE_META_GST; im=int(float(row.get("Impressions",0) or 0))
        except: continue
        if sp<=0 and im<=0: continue
        g_spend[cls]["spend"]+=sp; g_spend[cls]["impressions"]+=im; g_active.add(c)
    for cls in g_spend: g_spend[cls]["spend"]=round(g_spend[cls]["spend"],2)

    m_spend = defaultdict(esp); m_active = set()
    for row in meta_rows:
        c=row.get("Campaign","").strip(); ad=row.get("Ad","").strip()
        if not is_retention_meta(c) or not in_range(row.get("Date",""), ps, pe): continue
        cls=classify_meta_ad(ad)
        try: sp=float(row.get("Spend_INR",0) or 0)*GOOGLE_META_GST; im=int(float(row.get("Impressions",0) or 0))
        except: continue
        if sp<=0 and im<=0: continue
        m_spend[cls]["spend"]+=sp; m_spend[cls]["impressions"]+=im; m_active.add(c)
    for cls in m_spend: m_spend[cls]["spend"]=round(m_spend[cls]["spend"],2)

    g_reach = 0
    try:
        for row in unique_rows:
            c=row.get("Campaign","").strip()
            if c not in g_active: continue
            uu=row.get("Unique users","").strip()
            if not uu or uu=="--": continue
            try: g_reach+=int(str(uu).replace(",",""))
            except: continue
    except Exception as e: print(f"  [GReach] {e}")

    m_reach = 0
    at=os.getenv("META_ACCESS_TOKEN"); aa=os.getenv("META_AD_ACCOUNT_ID")
    if at and aa:
        try:
            url=f"https://graph.facebook.com/v18.0/{aa}/insights"
            params={"access_token":at,"level":"campaign","fields":"campaign_name,reach",
                    "time_range":f'{{"since":"{ps_str}","until":"{pe_str}"}}','limit':500}
            for item in req.get(url,params=params,timeout=30).json().get("data",[]):
                if item.get("campaign_name","") in m_active:
                    m_reach+=int(item.get("reach",0) or 0)
        except Exception as e: print(f"  [MetaReach] {e}")

    g_app = defaultdict(eap); m_app = defaultdict(eap)
    for row in apptrove_rows:
        partner=row.get("partner","").strip(); c=row.get("campaign","").strip()
        ag=row.get("ad_group","").strip(); ad=row.get("ad","").strip()
        channel=row.get("channel","").strip()
        if not in_range(row.get("Date",""), ps, pe): continue
        try: ao=int(row.get("app_opened",0) or 0); pu=int(row.get("purchase",0) or 0)
        except: ao=pu=0
        if partner=="Google Ads (Adwords)":
            if not channel or channel.strip()=="-": continue
            if not is_retention_google(c): continue
            cls=classify_google_adgroup(ag)
            g_app[cls]["app_opened"]+=ao; g_app[cls]["purchase"]+=pu
        elif partner=="Facebook":
            if not is_retention_meta(c): continue
            cls=classify_meta_ad(ad)
            m_app[cls]["app_opened"]+=ao; m_app[cls]["purchase"]+=pu

    # Aggregate to platform totals
    def platform_totals(spend, app):
        tot_im = sum(v["impressions"] for v in spend.values())
        tot_sp = round(sum(v["spend"] for v in spend.values()), 2)
        tot_ao = sum(v["app_opened"] for v in app.values())
        tot_pu = sum(v["purchase"]   for v in app.values())
        return tot_im, tot_sp, tot_ao, tot_pu

    m_im, m_sp, m_ao, m_pu = platform_totals(m_spend, m_app)
    g_im, g_sp, g_ao, g_pu = platform_totals(g_spend, g_app)

    print(f"  Meta  reach:{m_reach} impr:{m_im} spend:{m_sp} app:{m_ao} orders:{m_pu}")
    print(f"  Google reach:{g_reach} impr:{g_im} spend:{g_sp} app:{g_ao} orders:{g_pu}")

    h = ["Platform","Reach","Impressions","CTR (App/Impr%)","Spend","App Traffic","# Orders"]

    def make_row(platform, reach, im, sp, ao, pu):
        return [platform, reach, im, sdiv(ao, im), sp, ao, pu]

    meta_row   = make_row("Meta",   m_reach, m_im, m_sp, m_ao, m_pu)
    google_row = make_row("Google", g_reach, g_im, g_sp, g_ao, g_pu)

    # Total row — sum numeric columns
    tot_reach = m_reach + g_reach
    tot_im    = m_im + g_im
    tot_sp    = round(m_sp + g_sp, 2)
    tot_ao    = m_ao + g_ao
    tot_pu    = m_pu + g_pu
    total_row = make_row("Total", tot_reach, tot_im, tot_sp, tot_ao, tot_pu)

    return (
        [[f"3. RETENTION — {label}"], h,
         meta_row, google_row, total_row]
    )


# ── WRITE ──────────────────────────────────────────────────────────────────────
def write_sheet(gc, all_rows, ts):
    sh = gc.open_by_key(REPORT_SHEET_ID)
    try: ws = sh.worksheet(REPORT_TAB)
    except gspread.WorksheetNotFound: ws = sh.add_worksheet(title=REPORT_TAB, rows=300, cols=15)
    output = [[f"Previous Month Report — Last Updated: {ts}"], []] + all_rows
    ws.clear(); ws.update(output, value_input_option="USER_ENTERED")
    print(f"[Sheets] Written to '{REPORT_TAB}'")

def write_excel(all_rows, ts):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os as _os
    if not _os.path.exists(ONEDRIVE_PATH): print(f"[Excel] Not found — skipping"); return
    wb = load_workbook(ONEDRIVE_PATH)
    if REPORT_TAB in wb.sheetnames: del wb[REPORT_TAB]
    ws = wb.create_sheet(REPORT_TAB)
    sf=Font(bold=True,color="FFFFFF"); sx=PatternFill("solid",fgColor="2E75B6")
    pf=Font(bold=True,color="FFFFFF"); px=PatternFill("solid",fgColor="375623")
    hf=Font(bold=True,color="FFFFFF"); hx=PatternFill("solid",fgColor="1F4E79")
    PLATS = ("Google","Meta","WhatsApp","WhatsApp (Wati+WE)")
    cr = 1
    ws.cell(row=cr,column=1,value=f"Previous Month Report — Last Updated: {ts}").font=Font(italic=True); cr+=2
    for row in all_rows:
        if not row: cr+=1; continue
        first = str(row[0]) if row[0] else ""
        is_sec  = first.startswith(("1.","2.","3."))
        is_plat = first in PLATS
        is_note = first.startswith(("*","Note"))
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=cr, column=ci)
            try: cell.value = float(val) if val not in ("","N/A",None) else val
            except: cell.value = val
            if is_sec:   cell.font=sf; cell.fill=sx; cell.alignment=Alignment(horizontal="center")
            elif is_plat: cell.font=pf; cell.fill=px
            elif is_note: cell.font=Font(italic=True)
        cr += 1
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = ml+4
    wb.save(ONEDRIVE_PATH); print(f"[Excel] Written '{REPORT_TAB}' to OneDrive")


# ── MAIN ───────────────────────────────────────────────────────────────────────
def main():
    print("="*60)
    print("pull_prev_month_report.py — Previous Month Report")
    print("="*60)
    ps, pe, label = get_prev_month_range()
    print(f"[Period] {label}: {ps} → {pe}\n")
    gc  = get_gc(); ts = datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S")

    # Read each source tab ONCE — prevents Google Sheets 429 quota error
    print("[Data] Reading source tabs...")
    import time as _time
    google_rows   = read_sheet(gc, GOOGLE_SHEET_ID, "Google_Ads");        _time.sleep(2)
    meta_rows     = read_sheet(gc, GOOGLE_SHEET_ID, "Meta_Ads");          _time.sleep(2)
    wati_rows     = read_sheet(gc, GOOGLE_SHEET_ID, "Wati");              _time.sleep(2)
    apptrove_rows = read_sheet(gc, GOOGLE_SHEET_ID, "Apptrove_MMP");      _time.sleep(2)
    we_rows       = read_sheet(gc, WEBENGAGE_SHEET_ID, "Whatsapp_Campaign"); _time.sleep(2)
    try:
        unique_rows = read_sheet(gc, GOOGLE_UNIQUE_USERS_SID, GOOGLE_UNIQUE_USERS_TAB)
    except Exception as e:
        print(f"  [Unique Users] Error: {e}"); unique_rows = []
    print(f"  Loaded: G:{len(google_rows)} M:{len(meta_rows)} W:{len(wati_rows)} "
          f"App:{len(apptrove_rows)} WE:{len(we_rows)} UU:{len(unique_rows)}")

    all_rows = (
        build_cost_section(google_rows, meta_rows, wati_rows, we_rows, ps, pe, label)             + [[]] * 2 +
        build_onboarding_section(google_rows, meta_rows, wati_rows, we_rows, apptrove_rows,
                                 unique_rows, gc, ps, pe, label)                                   + [[]] * 2 +
        build_retention_section(google_rows, meta_rows, apptrove_rows, unique_rows, gc, ps, pe, label)
    )
    write_sheet(gc, all_rows, ts)
    write_excel(all_rows, ts)
    print(f"\n✅ Done — {label}")

if __name__ == "__main__":
    main()