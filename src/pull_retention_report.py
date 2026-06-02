"""
pull_retention_report.py
Builds the Retention Brand Report with Brand and Subcat rows for Google and Meta.

Three separate tables: MTD | MTD-1 | % Change

Reads from:
  - Original sheet (GOOGLE_SHEET_ID): Google_Ads, Meta_Ads, Apptrove_MMP
  - Google Unique Users sheet (GOOGLE_UNIQUE_USERS_SHEET_ID): campaign_unique tab

Writes to:
  - Report sheet (REPORT_SHEET_ID): retention_brand_report tab
  - OneDrive Excel: Team Wise Cost.xlsx -> retention_brand_report sheet

Reach logic:
  - MTD reach: fetched fresh from API/sheet
  - MTD-1 reach: read from previous run's MTD table in sheet
  - First run: MTD-1 reach = MTD reach
"""

import os
from datetime import datetime, timedelta
from collections import defaultdict
import pytz
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials
import requests as req

load_dotenv()

IST = pytz.timezone("Asia/Kolkata")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

GOOGLE_SHEET_ID         = os.getenv("GOOGLE_SHEET_ID")
REPORT_SHEET_ID         = os.getenv("REPORT_SHEET_ID")
GOOGLE_UNIQUE_USERS_SID = os.getenv("GOOGLE_UNIQUE_USERS_SHEET_ID")
GOOGLE_UNIQUE_USERS_TAB = os.getenv("GOOGLE_UNIQUE_USERS_TAB", "campaign_unique")

GOOGLE_META_GST = 1.18
ONEDRIVE_PATH = os.getenv("ONEDRIVE_EXCEL_PATH", r"C:\Users\anitb\OneDrive - Agrim Wholesale Private Limited\Marketing_Reports\Team Wise Cost.xlsx")

GOOGLE_RETENTION_KEYWORDS = ["ace", "ar_purchasers"]
META_RETENTION_KEYWORDS   = ["retention"]

HEADERS = [
    "KAM Campaigns", "Reach", "Impressions", "CTR (App/Impr%)",
    "Spend", "App Traffic", "# Orders"
]


# ── Auth ───────────────────────────────────────────────────────────────────────

def get_gc():
    creds_path = (
        os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH")
        or os.getenv("GOOGLE_CREDENTIALS_PATH")
        or "service_account.json"
    )
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


# ── Date periods ───────────────────────────────────────────────────────────────

def get_periods():
    now_ist   = datetime.now(IST)
    today     = now_ist.date()
    yesterday = today - timedelta(days=1)
    mtd_start = today.replace(day=1)

    if now_ist.hour < 12:
        mtd_end  = yesterday
        mtd1_end = yesterday - timedelta(days=1)
    else:
        mtd_end  = today
        mtd1_end = yesterday

    return mtd_start, mtd_end, mtd1_end


def in_range(date_val, start, end):
    if isinstance(date_val, str):
        for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
            try:
                d = datetime.strptime(date_val[:10], fmt).date()
                return start <= d <= end
            except ValueError:
                continue
        return False
    if hasattr(date_val, 'date'):
        return start <= date_val.date() <= end
    return start <= date_val <= end


# ── Classification ─────────────────────────────────────────────────────────────

def classify_google_adgroup(adgroup):
    n = adgroup.lower()
    if "brand" in n:
        return "Brand"
    if "categor" in n:
        return "Subcat"
    return None


def classify_meta_ad(ad):
    n = ad.lower()
    if "brand" in n:
        return "Brand"
    if "category" in n or "categor" in n:
        return "Subcat"
    if "catalogue" in n:
        return "Subcat"
    if "gibberellic" in n:
        return "Subcat"
    return "Brand"


def is_retention_google(campaign):
    n = campaign.lower()
    return any(k in n for k in GOOGLE_RETENTION_KEYWORDS)


def is_retention_meta(campaign):
    n = campaign.lower()
    return any(k in n for k in META_RETENTION_KEYWORDS)


# ── Read sheet ─────────────────────────────────────────────────────────────────

def read_sheet(gc, sheet_id, tab_name, retries=4, backoff=10):
    import time as _t
    for attempt in range(1, retries + 1):
        try:
            sh   = gc.open_by_key(sheet_id)
            ws   = sh.worksheet(tab_name)
            rows = ws.get_all_values()
            if not rows or len(rows) < 2:
                return []
            headers = rows[0]
            return [dict(zip(headers, row)) for row in rows[1:]]
        except Exception as e:
            msg = str(e)
            if attempt < retries and any(c in msg for c in ["503","500","429","502"]):
                wait = backoff * attempt
                print(f"  [read_sheet] {tab_name} attempt {attempt} failed, retrying in {wait}s...")
                _t.sleep(wait)
            else:
                raise


# ── Read previous reach from existing sheet ────────────────────────────────────

def read_previous_reach(gc):
    """
    Read MTD reach values from previous run's retention_brand_report tab.
    Returns dict: {platform: {cls: reach}}
    e.g. {"Google": {"Brand": 12345, "Subcat": 0}, "Meta": {"Brand": 67890, "Subcat": 0}}
    """
    result = {
        "Google": {"Brand": None, "Subcat": None},
        "Meta":   {"Brand": None, "Subcat": None},
    }
    try:
        sh = gc.open_by_key(REPORT_SHEET_ID)
        ws = sh.worksheet("retention_brand_report")
        all_rows = ws.get_all_values()

        # Find MTD table — look for row with "MTD" label
        in_mtd_table   = False
        current_platform = None

        for row in all_rows:
            if not any(row):
                in_mtd_table = False
                current_platform = None
                continue

            first = row[0].strip() if row[0] else ""

            if first == "MTD":
                in_mtd_table = True
                continue

            if in_mtd_table:
                if first in ("Meta", "Google"):
                    current_platform = first
                    continue
                if first in ("Brand", "Subcat") and current_platform:
                    try:
                        reach = int(str(row[1]).replace(",", "").replace(".0", "")) if row[1] else 0
                        result[current_platform][first] = reach
                    except (ValueError, IndexError):
                        pass
                # Stop at next table label
                if first in ("MTD-1", "% Change"):
                    break

    except Exception as e:
        print(f"[PrevReach] Could not read previous reach: {e}")

    print(f"[PrevReach] Google: {result['Google']} | Meta: {result['Meta']}")
    return result


# ── Process Google ─────────────────────────────────────────────────────────────

def process_google(gc, mtd_start, mtd_end, mtd1_end):
    print("[Google] Reading Google_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Google_Ads")

    def empty():
        return {"spend": 0.0, "impressions": 0}

    mtd  = defaultdict(empty)
    mtd1 = defaultdict(empty)
    adgroup_mtd = defaultdict(empty)   # adgroup-level for detail table
    active_campaigns = set()

    for row in rows:
        campaign = row.get("Campaign", "").strip()
        adgroup  = row.get("Ad_Group", "").strip()
        date_val = row.get("Date", "")

        if not is_retention_google(campaign):
            continue
        cls = classify_google_adgroup(adgroup)

        try:
            spend = float(row.get("Spend_INR", 0) or 0)
            impr  = int(float(row.get("Impressions", 0) or 0))
        except ValueError:
            continue

        if spend <= 0 and impr <= 0:
            continue

        spend_gst = spend * GOOGLE_META_GST

        if in_range(date_val, mtd_start, mtd_end):
            mtd[cls]["spend"]              += spend_gst
            mtd[cls]["impressions"]        += impr
            adgroup_mtd[adgroup]["spend"]       += spend_gst
            adgroup_mtd[adgroup]["impressions"] += impr
            active_campaigns.add(campaign)

        if in_range(date_val, mtd_start, mtd1_end):
            mtd1[cls]["spend"]       += spend_gst
            mtd1[cls]["impressions"] += impr

    for period in [mtd, mtd1, adgroup_mtd]:
        for k in period:
            period[k]["spend"] = round(period[k]["spend"], 2)

    return mtd, mtd1, active_campaigns, adgroup_mtd


def process_google_unique_users(gc, active_campaigns):
    print("[Google] Reading Unique Users sheet...")
    rows = read_sheet(gc, GOOGLE_UNIQUE_USERS_SID, GOOGLE_UNIQUE_USERS_TAB)
    total = 0
    for row in rows:
        campaign = row.get("Campaign", "").strip()
        uu_str   = row.get("Unique users", "").strip()
        if campaign not in active_campaigns:
            continue
        if not uu_str or uu_str == "--":
            continue
        try:
            total += int(str(uu_str).replace(",", ""))
        except ValueError:
            continue
    print(f"  Google unique users: {total}")
    return total


# ── Process Meta ───────────────────────────────────────────────────────────────

def process_meta(gc, mtd_start, mtd_end, mtd1_end):
    print("[Meta] Reading Meta_Ads...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Meta_Ads")

    def empty():
        return {"spend": 0.0, "impressions": 0}

    mtd  = defaultdict(empty)
    mtd1 = defaultdict(empty)
    ad_mtd = defaultdict(empty)   # ad-level for detail table
    active_campaigns = set()

    for row in rows:
        campaign = row.get("Campaign", "").strip()
        ad       = row.get("Ad", "").strip()
        date_val = row.get("Date", "")

        if not is_retention_meta(campaign):
            continue
        cls = classify_meta_ad(ad)

        try:
            spend = float(row.get("Spend_INR", 0) or 0)
            impr  = int(float(row.get("Impressions", 0) or 0))
        except ValueError:
            continue

        if spend <= 0 and impr <= 0:
            continue

        spend_gst = spend * GOOGLE_META_GST

        if in_range(date_val, mtd_start, mtd_end):
            mtd[cls]["spend"]        += spend_gst
            mtd[cls]["impressions"]  += impr
            ad_mtd[ad]["spend"]      += spend_gst
            ad_mtd[ad]["impressions"]+= impr
            active_campaigns.add(campaign)

        if in_range(date_val, mtd_start, mtd1_end):
            mtd1[cls]["spend"]       += spend_gst
            mtd1[cls]["impressions"] += impr

    for period in [mtd, mtd1, ad_mtd]:
        for k in period:
            period[k]["spend"] = round(period[k]["spend"], 2)

    return mtd, mtd1, active_campaigns, ad_mtd


def fetch_meta_mtd_reach(mtd_start, mtd_end, active_campaigns):
    access_token = os.getenv("META_ACCESS_TOKEN")
    ad_account   = os.getenv("META_AD_ACCOUNT_ID")
    if not access_token or not ad_account:
        return {"Brand": 0, "Subcat": 0}

    # We can't split reach by Brand/Subcat at campaign level
    # So total reach is for all retention campaigns combined
    url    = f"https://graph.facebook.com/v18.0/{ad_account}/insights"
    params = {
        "access_token": access_token,
        "level":        "campaign",
        "fields":       "campaign_name,reach",
        "time_range":   f'{{"since":"{mtd_start}","until":"{mtd_end}"}}',
        "limit":        500,
    }
    total = 0
    try:
        r    = req.get(url, params=params, timeout=30)
        data = r.json().get("data", [])
        for item in data:
            if item.get("campaign_name", "") in active_campaigns:
                total += int(item.get("reach", 0) or 0)
    except Exception as e:
        print(f"[Meta Reach] Error: {e}")
    print(f"[Meta] MTD reach: {total}")
    return total


# ── Process Apptrove ───────────────────────────────────────────────────────────

def process_apptrove(gc, mtd_start, mtd_end, mtd1_end):
    print("[Apptrove] Reading Apptrove_MMP...")
    rows = read_sheet(gc, GOOGLE_SHEET_ID, "Apptrove_MMP")

    def empty():
        return {"app_opened": 0, "purchase": 0}

    google_mtd  = defaultdict(empty)
    google_mtd1 = defaultdict(empty)
    meta_mtd    = defaultdict(empty)
    meta_mtd1   = defaultdict(empty)
    google_ag_mtd = defaultdict(empty)   # adgroup level
    meta_ad_mtd   = defaultdict(empty)   # ad level

    for row in rows:
        partner  = row.get("partner", "").strip()
        channel  = row.get("channel", "").strip()
        campaign = row.get("campaign", "").strip()
        adgroup  = row.get("ad_group", "").strip()
        ad       = row.get("ad", "").strip()
        date_val = row.get("Date", "")

        try:
            app_opened = int(row.get("app_opened", 0) or 0)
            purchase   = int(row.get("purchase", 0) or 0)
        except ValueError:
            app_opened, purchase = 0, 0

        if partner == "Google Ads (Adwords)":
            if not channel or channel.strip() == "-":
                continue
            if not is_retention_google(campaign):
                continue
            cls = classify_google_adgroup(adgroup)
            if in_range(date_val, mtd_start, mtd_end):
                google_mtd[cls]["app_opened"]          += app_opened
                google_mtd[cls]["purchase"]            += purchase
                google_ag_mtd[adgroup]["app_opened"]   += app_opened
                google_ag_mtd[adgroup]["purchase"]     += purchase
            if in_range(date_val, mtd_start, mtd1_end):
                google_mtd1[cls]["app_opened"] += app_opened
                google_mtd1[cls]["purchase"]   += purchase

        elif partner == "Facebook":
            if not is_retention_meta(campaign):
                continue
            cls = classify_meta_ad(ad)
            if in_range(date_val, mtd_start, mtd_end):
                meta_mtd[cls]["app_opened"]      += app_opened
                meta_mtd[cls]["purchase"]        += purchase
                meta_ad_mtd[ad]["app_opened"]    += app_opened
                meta_ad_mtd[ad]["purchase"]      += purchase
            if in_range(date_val, mtd_start, mtd1_end):
                meta_mtd1[cls]["app_opened"] += app_opened
                meta_mtd1[cls]["purchase"]   += purchase

    return google_mtd, google_mtd1, meta_mtd, meta_mtd1, google_ag_mtd, meta_ad_mtd


# ── Build tables ───────────────────────────────────────────────────────────────

DETAIL_ROWS = 10

def build_summary_table(g_spend, m_spend, g_app, m_app, g_reach, m_reach, label):
    """Table 1 — platform level: Meta + Google + Total."""
    h = ["Platform","Reach","Impressions","CTR (App/Impr%)","Spend","App Traffic","# Orders"]
    m_impr = sum(v["impressions"] for v in m_spend.values())
    m_sp   = round(sum(v["spend"] for v in m_spend.values()), 2)
    m_ao   = sum(v["app_opened"] for v in m_app.values())
    m_pu   = sum(v["purchase"]   for v in m_app.values())
    g_impr = sum(v["impressions"] for v in g_spend.values())
    g_sp   = round(sum(v["spend"] for v in g_spend.values()), 2)
    g_ao   = sum(v["app_opened"] for v in g_app.values())
    g_pu   = sum(v["purchase"]   for v in g_app.values())
    t_reach = m_reach + g_reach
    t_impr  = m_impr + g_impr
    t_sp    = round(m_sp + g_sp, 2)
    t_ao    = m_ao + g_ao
    t_pu    = m_pu + g_pu
    return [
        [f"Retention — {label}"], h,
        ["Meta",   m_reach, m_impr, safe_div(m_ao, m_impr), m_sp, m_ao, m_pu],
        ["Google", g_reach, g_impr, safe_div(g_ao, g_impr), g_sp, g_ao, g_pu],
        ["Total",  t_reach, t_impr, safe_div(t_ao, t_impr), t_sp, t_ao, t_pu],
    ]


def build_detail_table(title, name_col, spend_dict, app_dict):
    """Tables 2 & 3 — ad/adgroup level, top 10 by spend, fixed rows."""
    h = [name_col,"Reach","Impressions","CTR (App/Impr%)","Spend","App Traffic","# Orders"]
    combined = defaultdict(lambda: {"spend":0.,"impressions":0,"app_opened":0,"purchase":0})
    for name, sd in spend_dict.items():
        combined[name]["spend"]       += sd["spend"]
        combined[name]["impressions"] += sd["impressions"]
    for name, ad in app_dict.items():
        combined[name]["app_opened"] += ad["app_opened"]
        combined[name]["purchase"]   += ad["purchase"]
    sorted_rows = sorted(combined.items(), key=lambda x: x[1]["spend"], reverse=True)[:DETAIL_ROWS]
    rows = [[title], h]
    for name, d in sorted_rows:
        rows.append([name, "N/A", d["impressions"],
                     safe_div(d["app_opened"], d["impressions"]),
                     round(d["spend"], 2), d["app_opened"], d["purchase"]])
    # Pad to fixed DETAIL_ROWS
    while len(rows) - 2 < DETAIL_ROWS:
        rows.append(["","","","","","",""])
    return rows



def safe_div(a, b):
    if not b:
        return "N/A"
    return round(a / b * 100, 2)


def pct_change(mtd_val, mtd1_val):
    try:
        mtd_val  = float(mtd_val)
        mtd1_val = float(mtd1_val)
    except (ValueError, TypeError):
        return "N/A"
    if not mtd1_val:
        return "N/A"
    return round((mtd_val - mtd1_val) / mtd1_val * 100, 1)


def build_platform_rows(platform, spend_mtd, spend_mtd1,
                         app_mtd, app_mtd1,
                         reach_mtd, reach_mtd1):
    rows = []
    rows.append([platform, "", "", "", "", "", ""])  # platform label row
    rows.append(HEADERS)

    for cls in ["Brand", "Subcat"]:
        impr_mtd  = spend_mtd[cls]["impressions"]
        sp_mtd    = spend_mtd[cls]["spend"]
        ao_mtd    = app_mtd[cls]["app_opened"]
        pu_mtd    = app_mtd[cls]["purchase"]
        ctr_mtd   = safe_div(ao_mtd, impr_mtd)
        rows.append([cls, reach_mtd, impr_mtd, ctr_mtd, sp_mtd, ao_mtd, pu_mtd])

    return rows


def build_tables(
    google_mtd, google_mtd1,
    google_reach_mtd, google_reach_mtd1,
    meta_mtd, meta_mtd1,
    meta_reach_mtd, meta_reach_mtd1,
    google_app_mtd, google_app_mtd1,
    meta_app_mtd, meta_app_mtd1,
    google_ag_mtd, google_ag_app_mtd,
    meta_ad_mtd, meta_ad_app_mtd,
    mtd_start, mtd_end, mtd1_end
):
    def make_pct_table(platform, spend_mtd, spend_mtd1, app_mtd, app_mtd1,
                        reach_mtd, reach_mtd1):
        rows = []
        rows.append([platform, "", "", "", "", "", ""])
        rows.append(HEADERS)
        for cls in ["Brand", "Subcat"]:
            impr_mtd  = spend_mtd[cls]["impressions"]
            impr_mtd1 = spend_mtd1[cls]["impressions"]
            sp_mtd    = spend_mtd[cls]["spend"]
            sp_mtd1   = spend_mtd1[cls]["spend"]
            ao_mtd    = app_mtd[cls]["app_opened"]
            ao_mtd1   = app_mtd1[cls]["app_opened"]
            pu_mtd    = app_mtd[cls]["purchase"]
            pu_mtd1   = app_mtd1[cls]["purchase"]
            ctr_mtd   = safe_div(ao_mtd, impr_mtd)
            ctr_mtd1  = safe_div(ao_mtd1, impr_mtd1)
            rows.append([
                cls,
                pct_change(reach_mtd, reach_mtd1),
                pct_change(impr_mtd, impr_mtd1),
                pct_change(ctr_mtd, ctr_mtd1),
                pct_change(sp_mtd, sp_mtd1),
                pct_change(ao_mtd, ao_mtd1),
                pct_change(pu_mtd, pu_mtd1),
            ])
        return rows

    # MTD — Table 1: Summary
    mtd_label = f"MTD ({mtd_start} → {mtd_end})"
    summary   = build_summary_table(google_mtd, meta_mtd, google_app_mtd, meta_app_mtd,
                                     google_reach_mtd, meta_reach_mtd, mtd_label)
    # MTD — Table 2: Meta Ad level (top 10)
    meta_detail = build_detail_table(f"Meta Ad Performance — {mtd_label}",
                                      "Ad", meta_ad_mtd, meta_ad_app_mtd)
    # MTD — Table 3: Google Ad Group level (top 10)
    google_detail = build_detail_table(f"Google Ad Group Performance — {mtd_label}",
                                        "Ad Group", google_ag_mtd, google_ag_app_mtd)
    mtd_table = summary + [[]] + meta_detail + [[]] + google_detail

    # MTD-1 table
    def build_mtd1_rows(platform, spend_mtd1, app_mtd1, reach_mtd1):
        rows = []
        rows.append([platform, "", "", "", "", "", ""])
        rows.append(HEADERS)
        for cls in ["Brand", "Subcat"]:
            impr  = spend_mtd1[cls]["impressions"]
            sp    = spend_mtd1[cls]["spend"]
            ao    = app_mtd1[cls]["app_opened"]
            pu    = app_mtd1[cls]["purchase"]
            ctr   = safe_div(ao, impr)
            rows.append([cls, reach_mtd1, impr, ctr, sp, ao, pu])
        return rows

    mtd1_label  = [f"MTD-1 ({mtd_start} → {mtd1_end})"]
    mtd1_table  = [mtd1_label]
    mtd1_table += build_mtd1_rows("Meta",   meta_mtd1,   meta_app_mtd1,   meta_reach_mtd1)
    mtd1_table += [[]]
    mtd1_table += build_mtd1_rows("Google", google_mtd1, google_app_mtd1, google_reach_mtd1)

    # % Change table
    pct_label  = ["% Change"]
    pct_table  = [pct_label]
    pct_table += make_pct_table("Meta",   meta_mtd,   meta_mtd1,   meta_app_mtd,   meta_app_mtd1,   meta_reach_mtd,   meta_reach_mtd1)
    pct_table += [[]]
    pct_table += make_pct_table("Google", google_mtd, google_mtd1, google_app_mtd, google_app_mtd1, google_reach_mtd, google_reach_mtd1)

    return mtd_table, mtd1_table, pct_table


# ── Write Google Sheet ─────────────────────────────────────────────────────────

def write_report(gc, mtd_table, mtd1_table, pct_table, now_ist_str):
    sh = gc.open_by_key(REPORT_SHEET_ID)
    try:
        ws = sh.worksheet("retention_brand_report")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title="retention_brand_report", rows=100, cols=10)

    all_rows = [[f"Last Updated: {now_ist_str}"], []]
    all_rows += mtd_table
    all_rows += [[]]
    all_rows += mtd1_table
    all_rows += [[]]
    all_rows += pct_table

    ws.clear()
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print(f"[Sheets] Written retention_brand_report tab")


# ── Write OneDrive Excel ───────────────────────────────────────────────────────

def write_excel(mtd_table, mtd1_table, pct_table, now_ist_str):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os as _os

    if not _os.path.exists(ONEDRIVE_PATH):
        print(f"[Excel] File not found: {ONEDRIVE_PATH}")
        return

    wb = load_workbook(ONEDRIVE_PATH)
    if "retention_brand_report" in wb.sheetnames:
        del wb["retention_brand_report"]
    ws = wb.create_sheet("retention_brand_report")

    label_font  = Font(bold=True, color="FFFFFF")
    label_fill  = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center")

    current_row = 1
    ws.cell(row=current_row, column=1, value=f"Last Updated: {now_ist_str}").font = Font(italic=True)
    current_row += 2

    for table in [mtd_table, mtd1_table, pct_table]:
        for row_idx, row in enumerate(table):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                try:
                    cell.value = float(val) if val not in ("", None, "N/A") else val
                except (ValueError, TypeError):
                    cell.value = val
                if row_idx == 0:
                    cell.font = label_font
                    cell.fill = label_fill
                    cell.alignment = center
                elif row and row[0] in ("Meta", "Google"):
                    cell.font = header_font
                    cell.fill = header_fill
                elif row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
            current_row += 1
        current_row += 1

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(ONEDRIVE_PATH)
    print(f"[Excel] Written retention_brand_report to OneDrive")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("pull_retention_report.py — Retention Brand Report")
    print("=" * 55)

    mtd_start, mtd_end, mtd1_end = get_periods()
    print(f"[Period] MTD: {mtd_start} → {mtd_end} | MTD-1: {mtd_start} → {mtd1_end}")

    gc          = get_gc()
    now_ist_str = datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S")

    # Read previous reach BEFORE overwriting sheet
    prev_reach = read_previous_reach(gc)

    # Google
    google_mtd, google_mtd1, google_active, google_ag_mtd = process_google(gc, mtd_start, mtd_end, mtd1_end)
    google_reach_mtd = process_google_unique_users(gc, google_active)

    # MTD-1 reach from previous run (first run: same as MTD)
    google_reach_mtd1 = prev_reach["Google"]["Brand"] or google_reach_mtd

    # Meta
    meta_mtd, meta_mtd1, meta_active, meta_ad_mtd = process_meta(gc, mtd_start, mtd_end, mtd1_end)
    meta_reach_mtd = fetch_meta_mtd_reach(
        mtd_start.strftime("%Y-%m-%d"),
        mtd_end.strftime("%Y-%m-%d"),
        meta_active
    )
    meta_reach_mtd1 = prev_reach["Meta"]["Brand"] or meta_reach_mtd

    # Apptrove
    google_app_mtd, google_app_mtd1, meta_app_mtd, meta_app_mtd1, google_ag_app_mtd, meta_ad_app_mtd = process_apptrove(
        gc, mtd_start, mtd_end, mtd1_end
    )

    # Build tables
    mtd_table, mtd1_table, pct_table = build_tables(
        google_mtd, google_mtd1,
        google_reach_mtd, google_reach_mtd1,
        meta_mtd, meta_mtd1,
        meta_reach_mtd, meta_reach_mtd1,
        google_app_mtd, google_app_mtd1,
        meta_app_mtd, meta_app_mtd1,
        google_ag_mtd, google_ag_app_mtd,
        meta_ad_mtd, meta_ad_app_mtd,
        mtd_start, mtd_end, mtd1_end
    )

    # Write
    write_report(gc, mtd_table, mtd1_table, pct_table, now_ist_str)
    write_excel(mtd_table, mtd1_table, pct_table, now_ist_str)

    print("\n[OK] pull_retention_report.py complete")


if __name__ == "__main__":
    main()