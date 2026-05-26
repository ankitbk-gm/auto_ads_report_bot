"""
pull_inapp_subcat_report.py
Builds two In-app Promotion Reports:

1. Subcategory Promotion Report
   Filter : campaign name starts with 'Subcategory_Promotion_'
   Store  : Inapp_Subcat_Store  |  Report: Inapp_Subcat_Report

2. Sale Promotion Report
   Filter : campaign name starts with 'sale_promotion_'
   Name   : strip prefix + trailing date (_25May26) → e.g. SasteKaAdress_1
   Store  : Inapp_Sale_Store    |  Report: Inapp_Sale_Report

Both read from WebEngage sheet (Inapp_Campaign tab).
Both use persistent monthly store — one row per name per month.
Metrics: Unique Impressions, Unique Clicks, CTR
"""

import os
import re
import json
from datetime import datetime, timedelta
from collections import defaultdict
import pytz
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials

load_dotenv()

IST = pytz.timezone("Asia/Kolkata")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

WEBENGAGE_SHEET_ID = os.getenv("WEBENGAGE_SHEET_ID")
REPORT_SHEET_ID    = os.getenv("REPORT_SHEET_ID")
ONEDRIVE_PATH = os.getenv("ONEDRIVE_EXCEL_PATH", r"C:\Users\anitb\OneDrive - Agrim Wholesale Private Limited\Marketing_Reports\Team Wise Cost.xlsx")

STORE_TAB   = "Inapp_Subcat_Store"
REPORT_TAB  = "Inapp_Subcat_Report"

SALE_STORE_TAB   = "Inapp_Sale_Store"
SALE_REPORT_TAB  = "Inapp_Sale_Report"

STORE_HEADERS  = ["Subcategory", "Month", "Impressions", "Clicks", "Status", "Launch_Date"]
REPORT_HEADERS = ["Subcategory", "Status", "Launch Date", "Impressions",
                  "Unique Outbound Click", "CTR",
                  "Avg PPV Per Day (7d) Before", "Avg PPV Per Day After", "Diff"]

SALE_STORE_HEADERS  = ["Name", "Month", "Impressions", "Clicks", "Status", "Launch_Date"]
SALE_REPORT_HEADERS = ["Sale Promotion", "Status", "Launch Date",
                       "Impressions", "Unique Outbound Clicks", "CTR"]

# Manually seeded entries for campaigns that ran before automation
# Format: {subcategory: {month: {impressions, clicks, status, launch_date}}}
MANUAL_SEED = {
    "Mulch Film": {
        "2026-04": {"impressions": 16873, "clicks": 7444, "status": "Paused", "launch_date": "2026-04-25"},
    },
    "Carbendzim 50": {
        "2026-04": {"impressions": 13500, "clicks": 6652, "status": "Paused", "launch_date": "2026-04-26"},
    },
    "Spare Parts": {
        "2026-04": {"impressions": 14599, "clicks": 5492, "status": "Paused", "launch_date": "2026-04-26"},
    },
}

# ── Auth ───────────────────────────────────────────────────────────────────────

def get_gc():
    creds_path = (
        os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH")
        or os.getenv("GOOGLE_CREDENTIALS_PATH")
        or "service_account.json"
    )
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


# ── Read sheet ─────────────────────────────────────────────────────────────────

def read_sheet(gc, sheet_id, tab_name):
    sh   = gc.open_by_key(sheet_id)
    ws   = sh.worksheet(tab_name)
    rows = ws.get_all_values()
    if not rows or len(rows) < 2:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def get_or_create_tab(gc, sheet_id, tab_name, rows=1000, cols=20):
    sh = gc.open_by_key(sheet_id)
    try:
        return sh.worksheet(tab_name)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=tab_name, rows=rows, cols=cols)
        print(f"  [Sheets] Created tab '{tab_name}'")
        return ws


# ── Subcategory extraction ─────────────────────────────────────────────────────

# Trailing segment/screen/date pattern
TRAILING = re.compile(
    r'(_[Ss]eg[\d_,\s]+_full_Screen.+|_full_[Ss]creen.+|_seg.+)$',
    re.IGNORECASE
)


def extract_subcat(campaign_name):
    """Extract subcategory name from Subcategory_Promotion_* campaign name."""
    n = campaign_name.strip()
    # Must start with Subcategory_Promotion_
    m = re.match(r'Subcategory_Promotion__?(.+)', n, re.IGNORECASE)
    if not m:
        return None
    raw = m.group(1).strip()
    # Strip trailing segment/screen/date
    raw = TRAILING.sub('', raw).strip()
    # Replace underscores with spaces (only if no spaces — names with spaces are fine)
    if ' ' not in raw:
        raw = raw.replace('_', ' ')
    return raw.strip()


# Trailing date suffix pattern for sale promotions e.g. _25May26 _1May2026
SALE_DATE_SUFFIX = re.compile(r'_\d{1,2}[A-Za-z]{3}\d{2,4}$')


def extract_sale_name(campaign_name):
    """
    Extract sale promotion name from sale_promotion_* campaign name.
    Strips prefix and trailing date suffix, keeps version number.
    Example: sale_promotion_SasteKaAdress_1_25May26 → SasteKaAdress_1
    """
    n = campaign_name.strip()
    m = re.match(r'sale_promotion_(.+)', n, re.IGNORECASE)
    if not m:
        return None
    raw = m.group(1).strip()
    # Strip trailing date like _25May26
    raw = SALE_DATE_SUFFIX.sub('', raw).strip('_')
    return raw if raw else None


def parse_date(date_val):
    """Parse various date formats to date object."""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    if hasattr(date_val, 'date'):
        return date_val.date()
    s = str(date_val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


# ── Read persistent store ──────────────────────────────────────────────────────

def read_store(gc):
    """
    Read Inapp_Subcat_Store tab.
    Returns dict: {subcategory: {month: {impressions, clicks, status, launch_date}}}
    """
    store = defaultdict(dict)
    try:
        rows = read_sheet(gc, REPORT_SHEET_ID, STORE_TAB)
        for row in rows:
            subcat     = row.get("Subcategory", "").strip()
            month      = row.get("Month", "").strip()
            if not subcat or not month:
                continue
            store[subcat][month] = {
                "impressions":  int(float(row.get("Impressions", 0) or 0)),
                "clicks":       int(float(row.get("Clicks", 0) or 0)),
                "status":       row.get("Status", "").strip(),
                "launch_date":  row.get("Launch_Date", "").strip(),
            }
    except gspread.WorksheetNotFound:
        print(f"  [Store] Tab '{STORE_TAB}' not found — will create")
    return store


# ── Read WebEngage Inapp data ──────────────────────────────────────────────────

def read_inapp(gc):
    """
    Read Inapp_Campaign tab from WebEngage sheet.
    Returns list of dicts with subcategory, month, impressions, clicks, status, launch_date.
    """
    rows = read_sheet(gc, WEBENGAGE_SHEET_ID, "Inapp_Campaign")
    results = []

    for row in rows:
        campaign = row.get("Campaign Name", "").strip()
        subcat   = extract_subcat(campaign)
        if not subcat:
            continue

        # Get reporting period month from start date
        period_start = parse_date(row.get("Reporting Period Start Date", ""))
        if period_start is None:
            continue
        month = period_start.strftime("%Y-%m")

        # Launch date = campaign start date
        launch_date = parse_date(row.get("Campaign Start Date", ""))

        try:
            impressions = int(float(row.get("Unique Impressions", 0) or 0))
            clicks      = int(float(row.get("Unique Clicks", 0) or 0))
        except (ValueError, TypeError):
            impressions, clicks = 0, 0

        status = row.get("Status", "").strip()

        results.append({
            "subcategory":  subcat,
            "month":        month,
            "impressions":  impressions,
            "clicks":       clicks,
            "status":       status,
            "launch_date":  launch_date.strftime("%Y-%m-%d") if launch_date else "",
        })

    return results


# ── Sale Promotion functions ────────────────────────────────────────────────────

def read_sale_inapp(gc):
    """
    Read Inapp_Campaign tab and return sale_promotion_* campaigns.
    Returns list of dicts with name, month, impressions, clicks, status, launch_date.
    """
    rows = read_sheet(gc, WEBENGAGE_SHEET_ID, "Inapp_Campaign")
    results = []

    for row in rows:
        campaign = row.get("Campaign Name", "").strip()
        name     = extract_sale_name(campaign)
        if not name:
            continue

        period_start = parse_date(row.get("Reporting Period Start Date", ""))
        if period_start is None:
            continue
        month = period_start.strftime("%Y-%m")

        launch_date = parse_date(row.get("Campaign Start Date", ""))

        try:
            impressions = int(float(row.get("Unique Impressions", 0) or 0))
            clicks      = int(float(row.get("Unique Clicks", 0) or 0))
        except (ValueError, TypeError):
            impressions, clicks = 0, 0

        status = row.get("Status", "").strip()

        results.append({
            "name":        name,
            "month":       month,
            "impressions": impressions,
            "clicks":      clicks,
            "status":      status,
            "launch_date": launch_date.strftime("%Y-%m-%d") if launch_date else "",
        })

    return results


def read_sale_store(gc):
    """
    Read Inapp_Sale_Store tab.
    Returns dict: {name: {month: {impressions, clicks, status, launch_date}}}
    """
    store = defaultdict(dict)
    try:
        rows = read_sheet(gc, REPORT_SHEET_ID, SALE_STORE_TAB)
        for row in rows:
            name  = row.get("Name", "").strip()
            month = row.get("Month", "").strip()
            if not name or not month:
                continue
            store[name][month] = {
                "impressions": int(float(row.get("Impressions", 0) or 0)),
                "clicks":      int(float(row.get("Clicks", 0) or 0)),
                "status":      row.get("Status", "").strip(),
                "launch_date": row.get("Launch_Date", "").strip(),
            }
    except gspread.WorksheetNotFound:
        print(f"  [Sale Store] Tab '{SALE_STORE_TAB}' not found — will create")
    return store


def update_sale_store(store, sale_rows):
    """Upsert sale promotion data into store. One row per (name, month)."""
    we_data = defaultdict(lambda: {
        "impressions": 0, "clicks": 0,
        "status": "Ended", "launch_date": ""
    })

    for row in sale_rows:
        key = (row["name"], row["month"])
        we_data[key]["impressions"] += row["impressions"]
        we_data[key]["clicks"]      += row["clicks"]

        new_ld = row["launch_date"]
        cur_ld = we_data[key]["launch_date"]
        if new_ld and (not cur_ld or new_ld < cur_ld):
            we_data[key]["launch_date"] = new_ld

        if row["status"] == "Running":
            we_data[key]["status"] = "Running"
        elif we_data[key]["status"] != "Running":
            we_data[key]["status"] = row["status"]

    for (name, month), data in we_data.items():
        if name not in store:
            store[name] = {}
        store[name][month] = {
            "impressions": data["impressions"],
            "clicks":      data["clicks"],
            "status":      data["status"],
            "launch_date": data["launch_date"],
        }
        print(f"  [Sale Store] Updated: {name} | {month} | "
              f"Impr={data['impressions']} | Clicks={data['clicks']} | {data['status']}")

    return store


def write_sale_store(gc, store):
    """Write full sale store to Inapp_Sale_Store tab."""
    ws = get_or_create_tab(gc, REPORT_SHEET_ID, SALE_STORE_TAB)

    rows = [SALE_STORE_HEADERS]
    for name in sorted(store.keys()):
        for month in sorted(store[name].keys()):
            d = store[name][month]
            rows.append([name, month, d["impressions"], d["clicks"],
                         d["status"], d["launch_date"]])

    ws.clear()
    ws.update(rows, value_input_option="USER_ENTERED")
    print(f"  [Sale Store] Written {len(rows) - 1} rows to '{SALE_STORE_TAB}'")


def build_sale_report(store):
    """Aggregate sale store across all months per name. No PPV columns."""
    rows = []

    for name in sorted(store.keys()):
        months     = store[name]
        total_impr = sum(d["impressions"] for d in months.values())
        total_clk  = sum(d["clicks"]      for d in months.values())
        ctr        = round(total_clk / total_impr * 100, 2) if total_impr else "N/A"

        status = "Running" if any(d["status"] == "Running" for d in months.values()) else "Ended"

        launch_dates = [d["launch_date"] for d in months.values() if d["launch_date"]]
        launch_date  = min(launch_dates) if launch_dates else ""
        if launch_date:
            try:
                launch_date = datetime.strptime(launch_date, "%Y-%m-%d").strftime("%d-%m-%Y")
            except ValueError:
                pass

        rows.append([name, status, launch_date, total_impr, total_clk, ctr])

    # Running first, then by impressions desc
    rows.sort(key=lambda x: (0 if x[1] == "Running" else 1,
                              -x[3] if isinstance(x[3], int) else 0))
    return rows


def write_sale_report(gc, data_rows, now_ist_str):
    """Write sale promotion report to Inapp_Sale_Report tab."""
    sh = gc.open_by_key(REPORT_SHEET_ID)
    try:
        ws = sh.worksheet(SALE_REPORT_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=SALE_REPORT_TAB, rows=200, cols=10)

    all_rows = [
        [f"Last Updated: {now_ist_str}"],
        [],
        SALE_REPORT_HEADERS,
    ]
    all_rows.extend(data_rows)

    ws.clear()
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print(f"  [Sheets] Written {len(data_rows)} sale promotions to '{SALE_REPORT_TAB}'")


def write_sale_excel(data_rows, now_ist_str):
    """Write sale promotion report to Excel."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os as _os

    if not _os.path.exists(ONEDRIVE_PATH):
        print(f"[Excel] File not found: {ONEDRIVE_PATH}")
        return

    wb = load_workbook(ONEDRIVE_PATH)
    if SALE_REPORT_TAB in wb.sheetnames:
        del wb[SALE_REPORT_TAB]
    ws = wb.create_sheet(SALE_REPORT_TAB)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center")

    cr = 1
    ws.cell(row=cr, column=1, value=f"Last Updated: {now_ist_str}").font = Font(italic=True)
    cr += 2

    for ci, h in enumerate(SALE_REPORT_HEADERS, start=1):
        cell            = ws.cell(row=cr, column=ci, value=h)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center
    cr += 1

    for row in data_rows:
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=cr, column=ci)
            try:
                cell.value = float(val) if val not in ("", None, "N/A") else val
            except (ValueError, TypeError):
                cell.value = val
        cr += 1

    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = ml + 4

    wb.save(ONEDRIVE_PATH)
    print(f"  [Excel] Written '{SALE_REPORT_TAB}' to OneDrive")

def update_store(store, inapp_rows):
    """
    Upsert WebEngage data into store.
    One row per (subcategory, month) — always replace same month's data.
    Paused subcategories not in current data: untouched.
    """
    # Group WebEngage data by (subcat, month) — sum across segments
    we_data = defaultdict(lambda: {
        "impressions": 0, "clicks": 0,
        "status": "Ended", "launch_date": ""
    })

    for row in inapp_rows:
        key = (row["subcategory"], row["month"])
        we_data[key]["impressions"] += row["impressions"]
        we_data[key]["clicks"]      += row["clicks"]

        # Take oldest (minimum) launch date
        new_ld = row["launch_date"]
        cur_ld = we_data[key]["launch_date"]
        if new_ld:
            if not cur_ld or new_ld < cur_ld:
                we_data[key]["launch_date"] = new_ld
        # Running takes priority over Paused/Ended
        if row["status"] == "Running":
            we_data[key]["status"] = "Running"
        elif we_data[key]["status"] != "Running":
            we_data[key]["status"] = row["status"]

    # Upsert into store
    for (subcat, month), data in we_data.items():
        if subcat not in store:
            store[subcat] = {}
        store[subcat][month] = {
            "impressions":  data["impressions"],
            "clicks":       data["clicks"],
            "status":       data["status"],
            "launch_date":  data["launch_date"],
        }
        print(f"  [Store] Updated: {subcat} | {month} | "
              f"Impr={data['impressions']} | Clicks={data['clicks']} | {data['status']}")

    # Seed manual entries if not already in store
    for subcat, months in MANUAL_SEED.items():
        if subcat not in store:
            store[subcat] = {}
        for month, data in months.items():
            if month not in store[subcat]:
                store[subcat][month] = data
                print(f"  [Store] Seeded: {subcat} | {month}")

    return store


# ── Write store to sheet ───────────────────────────────────────────────────────

def write_store(gc, store):
    """Write full store to Inapp_Subcat_Store tab."""
    ws = get_or_create_tab(gc, REPORT_SHEET_ID, STORE_TAB)

    rows = [STORE_HEADERS]
    for subcat in sorted(store.keys()):
        for month in sorted(store[subcat].keys()):
            d = store[subcat][month]
            rows.append([
                subcat,
                month,
                d["impressions"],
                d["clicks"],
                d["status"],
                d["launch_date"],
            ])

    ws.clear()
    ws.update(rows, value_input_option="USER_ENTERED")
    print(f"  [Store] Written {len(rows) - 1} rows to '{STORE_TAB}'")


# ── Build report rows ──────────────────────────────────────────────────────────

def build_report(store):
    """
    Aggregate store across all months per subcategory.
    Returns list of report rows.
    """
    rows = []

    for subcat in sorted(store.keys()):
        months     = store[subcat]
        total_impr = sum(d["impressions"] for d in months.values())
        total_clk  = sum(d["clicks"]      for d in months.values())
        ctr        = round(total_clk / total_impr * 100, 2) if total_impr else "N/A"

        # Status: Running if any month has Running
        status = "Running" if any(d["status"] == "Running" for d in months.values()) else "Ended"

        # Launch date: earliest across months
        launch_dates = [d["launch_date"] for d in months.values() if d["launch_date"]]
        launch_date  = min(launch_dates) if launch_dates else ""
        # Format nicely
        if launch_date:
            try:
                launch_date = datetime.strptime(launch_date, "%Y-%m-%d").strftime("%d-%m-%Y")
            except ValueError:
                pass

        rows.append([
            subcat,
            status,
            launch_date,
            total_impr,
            total_clk,
            ctr,
            "",   # Avg PPV Per Day Before — manual
            "",   # Avg PPV Per Day After  — manual
            "",   # Diff                   — manual
        ])

    # Sort: Running first, then by total impressions desc
    rows.sort(key=lambda x: (0 if x[1] == "Running" else 1, -x[3] if isinstance(x[3], int) else 0))
    return rows


# ── Write report to Google Sheet ───────────────────────────────────────────────

def write_report(gc, data_rows, now_ist_str):
    sh = gc.open_by_key(REPORT_SHEET_ID)
    try:
        ws = sh.worksheet(REPORT_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=REPORT_TAB, rows=200, cols=15)

    all_rows = [
        [f"Last Updated: {now_ist_str}"],
        ["Note: PPV columns are filled manually"],
        [],
        REPORT_HEADERS,
    ]
    all_rows.extend(data_rows)

    ws.clear()
    ws.update(all_rows, value_input_option="USER_ENTERED")
    print(f"  [Sheets] Written {len(data_rows)} subcategories to '{REPORT_TAB}'")


# ── Write to OneDrive Excel ────────────────────────────────────────────────────

def write_excel(data_rows, now_ist_str):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os as _os

    if not _os.path.exists(ONEDRIVE_PATH):
        print(f"[Excel] File not found: {ONEDRIVE_PATH}")
        return

    wb = load_workbook(ONEDRIVE_PATH)
    if REPORT_TAB in wb.sheetnames:
        del wb[REPORT_TAB]
    ws = wb.create_sheet(REPORT_TAB)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center")

    cr = 1
    ws.cell(row=cr, column=1, value=f"Last Updated: {now_ist_str}").font = Font(italic=True); cr += 1
    ws.cell(row=cr, column=1, value="Note: PPV columns filled manually").font = Font(italic=True); cr += 2

    for ci, h in enumerate(REPORT_HEADERS, start=1):
        cell = ws.cell(row=cr, column=ci, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
    cr += 1

    for row in data_rows:
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=cr, column=ci)
            try:
                cell.value = float(val) if val not in ("", None, "N/A") else val
            except (ValueError, TypeError):
                cell.value = val
        cr += 1

    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = ml + 4

    wb.save(ONEDRIVE_PATH)
    print(f"  [Excel] Written '{REPORT_TAB}' to OneDrive")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("pull_inapp_subcat_report.py — In-app Subcat Report")
    print("=" * 55)

    gc          = get_gc()
    now_ist_str = datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S")

    # ── Subcategory Promotion ──────────────────────────────────────────────────
    print("\n[1] Reading persistent subcat store...")
    store = read_store(gc)
    print(f"  Store has {sum(len(v) for v in store.values())} entries "
          f"across {len(store)} subcategories")

    print("\n[2] Reading WebEngage Inapp_Campaign (subcategory)...")
    inapp_rows = read_inapp(gc)
    subcat_names = set(r["subcategory"] for r in inapp_rows)
    print(f"  Found {len(inapp_rows)} rows | {len(subcat_names)} subcategories: {subcat_names}")

    print("\n[3] Updating subcat store...")
    store = update_store(store, inapp_rows)

    print("\n[4] Writing subcat store...")
    write_store(gc, store)

    print("\n[5] Building subcat report...")
    data_rows = build_report(store)
    print(f"  {len(data_rows)} subcategories in report")

    print("\n[6] Writing subcat report...")
    write_report(gc, data_rows, now_ist_str)
    write_excel(data_rows, now_ist_str)

    # ── Sale Promotion ─────────────────────────────────────────────────────────
    print("\n" + "=" * 55)
    print("Sale Promotion Report")
    print("=" * 55)

    print("\n[7] Reading persistent sale store...")
    sale_store = read_sale_store(gc)
    print(f"  Store has {sum(len(v) for v in sale_store.values())} entries "
          f"across {len(sale_store)} promotions")

    print("\n[8] Reading WebEngage Inapp_Campaign (sale promotions)...")
    sale_rows = read_sale_inapp(gc)
    sale_names = set(r["name"] for r in sale_rows)
    print(f"  Found {len(sale_rows)} rows | {len(sale_names)} promotions: {sale_names}")

    print("\n[9] Updating sale store...")
    sale_store = update_sale_store(sale_store, sale_rows)

    print("\n[10] Writing sale store...")
    write_sale_store(gc, sale_store)

    print("\n[11] Building sale report...")
    sale_data_rows = build_sale_report(sale_store)
    print(f"  {len(sale_data_rows)} promotions in report")

    print("\n[12] Writing sale report...")
    write_sale_report(gc, sale_data_rows, now_ist_str)
    write_sale_excel(sale_data_rows, now_ist_str)

    print(f"\n✅ Complete — {len(data_rows)} subcategories | {len(sale_data_rows)} sale promotions")

    print("\n[Subcategory Summary]")
    print(f"  {'Subcategory':<50} {'Status':<10} {'Impressions':>12} {'Clicks':>8} {'CTR':>8}")
    print("  " + "-" * 95)
    for row in data_rows:
        print(f"  {row[0]:<50} {row[1]:<10} {row[3]:>12} {row[4]:>8} {str(row[5]):>8}")

    print("\n[Sale Promotion Summary]")
    print(f"  {'Name':<40} {'Status':<10} {'Impressions':>12} {'Clicks':>8} {'CTR':>8}")
    print("  " + "-" * 80)
    for row in sale_data_rows:
        print(f"  {row[0]:<40} {row[1]:<10} {row[3]:>12} {row[4]:>8} {str(row[5]):>8}")


if __name__ == "__main__":
    main()